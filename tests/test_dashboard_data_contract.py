from __future__ import annotations

import csv
import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "stock-selection-agent" / "scripts" / "build_dashboard_data.py"
SPEC = importlib.util.spec_from_file_location("build_dashboard_data", SCRIPT)
dashboard = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = dashboard
SPEC.loader.exec_module(dashboard)


class DashboardDataContractTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.base = Path(self.tmp.name)
        self.scores_dir = self.base / "outputs"
        self.output_dir = self.base / "data" / "dashboard"
        self.workbook = self.base / "data" / "validation" / "output" / "stock_selection_log.xlsx"
        self.scores_dir.mkdir(parents=True)
        self.workbook.parent.mkdir(parents=True)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def write_csv(self, path: Path, rows: list[dict[str, object]]) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    def write_sheet(self, workbook: Workbook, name: str, rows: list[dict[str, object]]) -> None:
        sheet = workbook.create_sheet(name)
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

    def make_validation_workbook(self) -> None:
        workbook = Workbook()
        del workbook[workbook.sheetnames[0]]
        self.write_sheet(
            workbook,
            "selection_runs",
            [
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "selection_time": "15:10:00",
                    "strategy_version": "v1_0",
                    "market_env": "震荡",
                    "total_selected_count": 2,
                    "data_source": "scores=outputs\\reviewed_scores.csv; candidates=data\\snapshots\\20260624_candidates.csv",
                    "operator": "test",
                    "notes": "",
                    "created_at": "2026-06-24 15:10:00",
                }
            ],
        )
        self.write_sheet(
            workbook,
            "selected_stocks",
            [
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "selection_time": "15:10:00",
                    "stock_code": "300001.SZ",
                    "stock_name": "测试科技",
                    "sector": "AI应用",
                    "industry": "AI应用",
                    "concept": "",
                    "strategy_label": "平台突破买点",
                    "participation_level": "轻仓试错",
                    "suggested_action": "轻仓试错",
                    "selection_price": 10,
                    "selection_close": 10,
                    "selection_reason": "平台突破",
                    "stop_loss_price": 9.7,
                    "take_profit_price": 10.5,
                    "risk_note": "市场量能不足",
                    "total_score": 72,
                    "rank_in_run": 1,
                    "price_missing": False,
                    "created_at": "2026-06-24 15:10:00",
                },
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "selection_time": "15:10:00",
                    "stock_code": "600002.SH",
                    "stock_name": "测试电力",
                    "sector": "电网设备",
                    "industry": "电网设备",
                    "concept": "",
                    "strategy_label": "趋势回踩买点",
                    "participation_level": "只观察",
                    "suggested_action": "只观察",
                    "selection_price": 20,
                    "selection_close": 20,
                    "selection_reason": "趋势回踩",
                    "stop_loss_price": 19.4,
                    "take_profit_price": 21,
                    "risk_note": "",
                    "total_score": 61,
                    "rank_in_run": 2,
                    "price_missing": False,
                    "created_at": "2026-06-24 15:10:00",
                },
            ],
        )
        self.write_sheet(
            workbook,
            "performance",
            [
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "300001.SZ",
                    "stock_name": "测试科技",
                    "sector": "AI应用",
                    "industry": "AI应用",
                    "concept": "",
                    "strategy_label": "平台突破买点",
                    "participation_level": "轻仓试错",
                    "total_score": 72,
                    "rank_in_run": 1,
                    "selection_price": 10,
                    "latest_price": 10.8,
                    "latest_price_date": "2026-06-29",
                    "return_latest_pct": 8,
                    "return_t1_close_pct": 2,
                    "return_t2_close_pct": 4,
                    "return_t3_close_pct": 8,
                    "return_t5_close_pct": "",
                    "return_t10_close_pct": "",
                    "max_gain_3d_pct": 9,
                    "max_drawdown_3d_pct": -1,
                    "hit_stop_loss": False,
                    "hit_take_profit": True,
                    "result_label": "成功",
                    "failure_reason": "",
                    "data_status": "ok",
                    "updated_at": "2026-06-29 16:00:00",
                },
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "600002.SH",
                    "stock_name": "测试电力",
                    "sector": "电网设备",
                    "industry": "电网设备",
                    "concept": "",
                    "strategy_label": "趋势回踩买点",
                    "participation_level": "只观察",
                    "total_score": 61,
                    "rank_in_run": 2,
                    "selection_price": 20,
                    "latest_price": 19,
                    "latest_price_date": "2026-06-29",
                    "return_latest_pct": -5,
                    "return_t1_close_pct": -1,
                    "return_t2_close_pct": -3,
                    "return_t3_close_pct": -5,
                    "return_t5_close_pct": "",
                    "return_t10_close_pct": "",
                    "max_gain_3d_pct": 1,
                    "max_drawdown_3d_pct": -6,
                    "hit_stop_loss": True,
                    "hit_take_profit": False,
                    "result_label": "失败",
                    "failure_reason": "T3亏损",
                    "data_status": "ok",
                    "updated_at": "2026-06-29 16:00:00",
                },
            ],
        )
        self.write_sheet(
            workbook,
            "future_prices",
            [
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "300001.SZ",
                    "stock_name": "????",
                    "trading_day_offset": "T1",
                    "price_date": "2026-06-25",
                    "open": 10.1,
                    "high": 10.4,
                    "low": 10.0,
                    "close": 10.2,
                    "volume": 100000,
                    "amount": 1020000,
                    "turnover_rate": 2.1,
                    "is_suspended": False,
                    "data_source": "unit",
                    "updated_at": "2026-06-25 16:00:00",
                },
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "300001.SZ",
                    "stock_name": "????",
                    "trading_day_offset": "T2",
                    "price_date": "2026-06-26",
                    "open": 10.3,
                    "high": 10.6,
                    "low": 10.2,
                    "close": 10.4,
                    "volume": 100000,
                    "amount": 1040000,
                    "turnover_rate": 2.2,
                    "is_suspended": False,
                    "data_source": "unit",
                    "updated_at": "2026-06-26 16:00:00",
                },
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "300001.SZ",
                    "stock_name": "????",
                    "trading_day_offset": "T3",
                    "price_date": "2026-06-29",
                    "open": 10.5,
                    "high": 10.9,
                    "low": 10.4,
                    "close": 10.8,
                    "volume": 100000,
                    "amount": 1080000,
                    "turnover_rate": 2.3,
                    "is_suspended": False,
                    "data_source": "unit",
                    "updated_at": "2026-06-29 16:00:00",
                },
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "stock_code": "600002.SH",
                    "stock_name": "????",
                    "trading_day_offset": "T1",
                    "price_date": "2026-06-25",
                    "open": 19.9,
                    "high": 20.1,
                    "low": 19.7,
                    "close": 19.8,
                    "volume": 100000,
                    "amount": 1980000,
                    "turnover_rate": 1.2,
                    "is_suspended": False,
                    "data_source": "unit",
                    "updated_at": "2026-06-25 16:00:00",
                },
            ],
        )
        self.write_sheet(
            workbook,
            "summary_by_run",
            [
                {
                    "run_id": "20260624_reviewed_v1_0",
                    "selection_date": "2026-06-24",
                    "strategy_version": "v1_0",
                    "market_env": "震荡",
                    "total_selected_count": 2,
                    "valid_stock_count": 2,
                    "win_rate_t1": 50,
                    "win_rate_t2": 50,
                    "win_rate_t3": 50,
                    "win_rate_t5": "",
                    "avg_return_t1_pct": 0.5,
                    "avg_return_t2_pct": 0.5,
                    "avg_return_t3_pct": 1.5,
                    "avg_return_t5_pct": "",
                    "avg_max_drawdown_3d_pct": -3.5,
                    "hit_stop_loss_count": 1,
                    "hit_take_profit_count": 1,
                    "best_stock": "测试科技",
                    "worst_stock": "测试电力",
                    "conclusion": "样例结论",
                    "updated_at": "2026-06-29 16:00:00",
                }
            ],
        )
        workbook.save(self.workbook)

    def test_builds_index_and_daily_payloads_without_network(self) -> None:
        self.make_validation_workbook()
        self.write_csv(
            self.scores_dir / "selection_scores_20260625.csv",
            [
                {
                    "rank": 1,
                    "symbol": "688001",
                    "name": "未复盘科技",
                    "sector": "半导体",
                    "total_score": 81,
                    "trend_score": 40,
                    "startup_score": 25,
                    "sector_score": 12,
                    "market_score": 4,
                    "decision": "强参与",
                    "continuation": "强延续",
                    "buy_model": "强势震荡再启动",
                    "notes": "趋势清晰",
                    "risks": "",
                    "hard_rejects": "",
                    "plan": "小仓试错",
                }
            ],
        )

        index = dashboard.build_dashboard(
            scores_dir=self.scores_dir,
            validation_workbook=self.workbook,
            output_dir=self.output_dir,
            project_root=self.base,
        )

        self.assertEqual(["20260625", "20260624"], [row["date"] for row in index["runs"]])
        self.assertTrue((self.output_dir / "runs_index.json").exists())
        self.assertTrue((self.output_dir / "runs" / "20260624.json").exists())
        self.assertTrue((self.output_dir / "runs" / "20260625.json").exists())

        reviewed = json.loads((self.output_dir / "runs" / "20260624.json").read_text(encoding="utf-8"))
        self.assertEqual(1, reviewed["schema_version"])
        self.assertEqual("20260624_reviewed_v1_0", reviewed["active_run_id"])
        self.assertEqual("ready", reviewed["review"]["status"])
        self.assertEqual(2, reviewed["metrics"]["selected_count"])
        self.assertEqual(50, reviewed["metrics"]["review"]["win_rate_t3"])
        self.assertEqual({"value": "AI应用", "count": 1}, reviewed["filters"]["sectors"][0])
        self.assertEqual("300001", reviewed["picks"][0]["symbol"])
        self.assertEqual(8, reviewed["picks"][0]["review"]["returns"]["return_t3_close_pct"])
        price_points = reviewed["picks"][0]["review"]["price_points"]
        self.assertEqual(["T1", "T2", "T3"], [row["trading_day_offset"] for row in price_points])
        self.assertEqual("2026-06-25", price_points[0]["price_date"])
        self.assertEqual(10.2, price_points[0]["close"])
        self.assertEqual(2.0, price_points[0]["return_pct"])
        self.assertEqual(8.0, price_points[2]["return_pct"])
        self.assertEqual(1, reviewed["strategy_effectiveness"]["reviewed_run_count"])
        self.assertEqual(2, reviewed["strategy_effectiveness"]["valid_stock_count"])
        self.assertEqual(1.5, reviewed["strategy_effectiveness"]["avg_return_t3_pct"])

        missing = json.loads((self.output_dir / "runs" / "20260625.json").read_text(encoding="utf-8"))
        self.assertEqual("missing_review", missing["review"]["status"])
        self.assertFalse(missing["review"]["has_review"])
        self.assertEqual("暂无复盘数据", missing["review"]["empty_state"]["title"])
        self.assertEqual(81, missing["metrics"]["top_score"])
        self.assertEqual("正向验证", missing["strategy_effectiveness"]["conclusion"])

    def test_react_frontend_serves_local_dashboard_data_in_dev(self) -> None:
        config = (ROOT / "frontend" / "vite.config.js").read_text(encoding="utf-8")

        self.assertIn("local-dashboard-data", config)
        self.assertIn("/data/dashboard/", config)
        self.assertIn("dashboardDataRoot", config)
        self.assertIn("@vitejs/plugin-react", config)
        self.assertNotIn("service_role", config.lower())


if __name__ == "__main__":
    unittest.main()
