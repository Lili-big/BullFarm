from __future__ import annotations

import argparse
import csv
import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "stock-selection-agent" / "scripts" / "sync_supabase.py"
SPEC = importlib.util.spec_from_file_location("sync_supabase", SCRIPT)
sync_supabase = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = sync_supabase
SPEC.loader.exec_module(sync_supabase)


class SupabasePayloadTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.base = Path(self.tmp.name)
        self.scores = self.base / "selection_scores.csv"
        self.candidates = self.base / "20260623_candidates.csv"
        self.metadata = self.base / "20260623_fetch_meta.json"
        self.report = self.base / "selection_report.md"
        self.workbook = self.base / "stock_selection_log.xlsx"
        self.dashboard_summary = self.base / "dashboard_summary.csv"
        self.write_csv(
            self.scores,
            [
                {
                    "rank": "1",
                    "symbol": "300223",
                    "name": "Sample Tech",
                    "sector": "Electronics",
                    "total_score": "70",
                    "trend_score": "40",
                    "startup_score": "25",
                    "sector_score": "5",
                    "market_score": "0",
                    "decision": "watch",
                    "continuation": "weak",
                    "buy_model": "breakout",
                    "notes": "structure clear",
                    "risks": "sector weak",
                    "hard_rejects": "",
                    "plan": "watch pool",
                }
            ],
        )
        self.write_csv(
            self.candidates,
            [
                {
                    "symbol": "300223",
                    "name": "Sample Tech",
                    "sector": "Electronics",
                    "close": "100.5",
                    "ma20": "98.2",
                }
            ],
        )
        self.metadata.write_text(
            """{
              "generated_at": "2026-06-23T15:30:00",
              "end_date": "20260623",
              "source": "unit-test",
              "market": {"index_pct3": 1.2}
            }""",
            encoding="utf-8",
        )
        self.report.write_text("# report\n\npayload check", encoding="utf-8")
        self.write_csv(
            self.dashboard_summary,
            [
                {
                    "decision": "watch",
                    "sample_count": "1",
                    "t2_success_rate_pct": "100.0",
                }
            ],
        )
        self.write_workbook()

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def write_csv(self, path: Path, rows: list[dict[str, object]]) -> None:
        fields = list(rows[0].keys())
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)

    def write_workbook(self) -> None:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "selection_runs"
        ws.append(
            [
                "run_id",
                "selection_date",
                "selection_time",
                "strategy_version",
                "market_env",
                "total_selected_count",
                "data_source",
                "operator",
                "notes",
                "created_at",
            ]
        )
        ws.append(
            [
                "20260623_153000_v1_0",
                "2026-06-23",
                "15:30:00",
                "v1_0",
                "range",
                1,
                "scores=test; candidates=test",
                "unit",
                "note",
                "2026-06-23T15:30:00",
            ]
        )
        selected = workbook.create_sheet("selected_stocks")
        selected.append(["run_id", "stock_code", "stock_name", "sector", "total_score", "rank_in_run"])
        selected.append(["20260623_153000_v1_0", "300223.SZ", "Sample Tech", "Electronics", 70, 1])
        prices = workbook.create_sheet("future_prices")
        prices.append(
            [
                "run_id",
                "selection_date",
                "stock_code",
                "stock_name",
                "trading_day_offset",
                "price_date",
                "open",
                "high",
                "low",
                "close",
                "volume",
                "amount",
                "turnover_rate",
                "is_suspended",
                "data_source",
                "updated_at",
            ]
        )
        prices.append(
            [
                "20260623_153000_v1_0",
                "2026-06-23",
                "300223.SZ",
                "Sample Tech",
                "T1",
                "2026-06-24",
                100,
                105,
                99,
                104,
                1000000,
                100000000,
                2.1,
                False,
                "unit",
                "2026-06-24T15:00:00",
            ]
        )
        performance = workbook.create_sheet("performance")
        performance.append(
            [
                "run_id",
                "selection_date",
                "stock_code",
                "stock_name",
                "sector",
                "strategy_label",
                "participation_level",
                "total_score",
                "rank_in_run",
                "selection_price",
                "latest_price",
                "latest_price_date",
                "return_latest_pct",
                "return_t1_close_pct",
                "return_t2_close_pct",
                "return_t3_close_pct",
                "hit_stop_loss",
                "hit_take_profit",
                "is_profitable_t1",
                "result_label",
                "data_status",
                "updated_at",
            ]
        )
        performance.append(
            [
                "20260623_153000_v1_0",
                "2026-06-23",
                "300223.SZ",
                "Sample Tech",
                "Electronics",
                "breakout",
                "watch",
                70,
                1,
                100,
                104,
                "2026-06-24",
                4,
                4,
                None,
                None,
                False,
                True,
                True,
                "success",
                "partial",
                "2026-06-24T15:00:00",
            ]
        )
        summary = workbook.create_sheet("summary_by_run")
        summary.append(["run_id", "selection_date", "win_rate_t1", "avg_return_t1_pct"])
        summary.append(["20260623_153000_v1_0", "2026-06-23", 100.0, 4.0])
        for sheet_name, row in [
            ("selection_runs", ["20260622_153000_v1_0", "2026-06-22", "15:30:00", "v1_0", "range", 1, "scores=old", "unit", "old", "2026-06-22T15:30:00"]),
            ("selected_stocks", ["20260622_153000_v1_0", "600000.SH", "Old Bank", "Finance", 66, 1]),
            ("future_prices", ["20260622_153000_v1_0", "2026-06-22", "600000.SH", "Old Bank", "LATEST", "2026-06-24", 10, 11, 9, 10.5, 1000000, 10000000, 1.2, False, "unit", "2026-06-24T15:00:00"]),
            ("performance", ["20260622_153000_v1_0", "2026-06-22", "600000.SH", "Old Bank", "Finance", "breakout", "watch", 66, 1, 10, 10.5, "2026-06-24", 5, None, None, None, False, True, True, "success", "latest", "2026-06-24T15:00:00"]),
            ("summary_by_run", ["20260622_153000_v1_0", "2026-06-22", 100.0, 5.0]),
        ]:
            workbook[sheet_name].append(row)
        workbook.save(self.workbook)

    def make_args(self, **overrides: object) -> argparse.Namespace:
        values = {
            "scores": str(self.scores),
            "candidates": str(self.candidates),
            "metadata": str(self.metadata),
            "report": str(self.report),
            "workbook": str(self.workbook),
            "dashboard_summary": str(self.dashboard_summary),
            "dashboard_detail": None,
            "run_id": "20260623_153000_v1_0",
            "selection_date": None,
            "strategy_version": "v1_0",
            "market_env": None,
            "operator": None,
            "notes": None,
            "include_workbook_runs": False,
        }
        values.update(overrides)
        return argparse.Namespace(**values)

    def test_build_payload_maps_csv_metadata_dashboard_and_workbook(self) -> None:
        payload = sync_supabase.build_sync_payload(self.make_args())

        self.assertEqual(
            {
                "stock_selection_runs": 1,
                "stock_selection_results": 1,
                "stock_selection_prices": 1,
                "stock_selection_performance": 1,
            },
            payload.counts(),
        )
        run = payload.runs[0]
        result = payload.results[0]
        price = payload.prices[0]
        performance = payload.performance[0]
        self.assertEqual("20260623_153000_v1_0", run["run_id"])
        self.assertEqual("2026-06-23", run["selection_date"])
        self.assertEqual("unit-test", run["source_metadata"]["source"])
        self.assertEqual("unit-test", run["source_provider"])
        self.assertEqual("# report\n\npayload check", run["report_markdown"])
        self.assertEqual("watch", run["dashboard"]["dashboard_summary"][0]["decision"])
        self.assertEqual("300223.SZ", result["stock_code"])
        self.assertEqual(70, result["total_score"])
        self.assertEqual(100.5, result["selection_price"])
        self.assertEqual(100.5, result["candidate_payload"]["close"])
        self.assertEqual("T1", price["trading_day_offset"])
        self.assertEqual(104.0, price["close"])
        self.assertTrue(performance["hit_take_profit"])
        self.assertEqual(4.0, performance["return_t1_close_pct"])

    def test_include_workbook_runs_syncs_historical_price_updates(self) -> None:
        payload = sync_supabase.build_sync_payload(self.make_args(include_workbook_runs=True))

        self.assertEqual(
            {
                "stock_selection_runs": 2,
                "stock_selection_results": 2,
                "stock_selection_prices": 2,
                "stock_selection_performance": 2,
            },
            payload.counts(),
        )
        run_ids = {row["run_id"] for row in payload.runs}
        self.assertEqual({"20260623_153000_v1_0", "20260622_153000_v1_0"}, run_ids)
        old_result = next(row for row in payload.results if row["run_id"] == "20260622_153000_v1_0")
        old_price = next(row for row in payload.prices if row["run_id"] == "20260622_153000_v1_0")
        self.assertEqual("600000.SH", old_result["stock_code"])
        self.assertEqual("LATEST", old_price["trading_day_offset"])

    def test_conflict_keys_and_dry_run_skip_real_write(self) -> None:
        payload = sync_supabase.build_sync_payload(self.make_args())

        self.assertEqual("run_id", sync_supabase.UPSERT_CONFLICTS["stock_selection_runs"])
        self.assertEqual("run_id,stock_code", sync_supabase.UPSERT_CONFLICTS["stock_selection_results"])
        self.assertEqual("run_id,stock_code,trading_day_offset", sync_supabase.UPSERT_CONFLICTS["stock_selection_prices"])
        self.assertEqual("run_id,stock_code", sync_supabase.UPSERT_CONFLICTS["stock_selection_performance"])
        dry = sync_supabase.execute_sync(payload, dry_run=True, env={})
        missing = sync_supabase.execute_sync(payload, dry_run=False, env={})
        self.assertEqual("dry_run", dry["status"])
        self.assertEqual("skipped", missing["status"])
        self.assertIn("SUPABASE_URL", missing["reason"])

    def test_local_env_file_supplies_missing_supabase_credentials(self) -> None:
        env_file = self.base / "local.env"
        env_file.write_text(
            "SUPABASE_URL=https://example.supabase.co\n"
            "SUPABASE_SERVICE_ROLE_KEY=file-secret\n",
            encoding="utf-8",
        )

        loaded = sync_supabase.load_sync_env(env_file, base_env={"SUPABASE_SERVICE_ROLE_KEY": "system-secret"})

        self.assertEqual("https://example.supabase.co", loaded["SUPABASE_URL"])
        self.assertEqual("system-secret", loaded["SUPABASE_SERVICE_ROLE_KEY"])

    def test_emit_sql_builds_idempotent_upsert_chunk(self) -> None:
        payload = sync_supabase.build_sync_payload(self.make_args())
        sql = sync_supabase.build_upsert_sql("stock_selection_results", payload.results)

        self.assertIn("jsonb_populate_recordset", sql)
        self.assertIn('public."stock_selection_results"', sql)
        self.assertIn('on conflict ("run_id", "stock_code") do update', sql)
        self.assertIn('"selection_price"', sql)
        self.assertIn("Sample Tech", sql)

        compact_run_sql = sync_supabase.build_upsert_sql("stock_selection_runs", payload.runs, compact=True)
        self.assertNotIn("report_markdown", compact_run_sql)
        self.assertNotIn('"dashboard"', compact_run_sql)

    def test_sql_bundle_manifest_lists_chunks_for_plugin_handoff(self) -> None:
        payload = sync_supabase.build_sync_payload(self.make_args())
        bundle_dir = self.base / "sql_bundle"

        manifest = sync_supabase.write_sql_bundle(payload, bundle_dir, chunk_size=1, compact=True)

        self.assertTrue((bundle_dir / "manifest.json").exists())
        self.assertEqual(4, len(manifest["files"]))
        self.assertEqual(1, manifest["counts"]["stock_selection_runs"])
        self.assertTrue(any(item["table"] == "stock_selection_results" for item in manifest["files"]))
        first_sql = Path(manifest["files"][0]["path"]).read_text(encoding="utf-8")
        self.assertIn("on conflict", first_sql)

    def test_migration_contains_supabase_security_contract(self) -> None:
        migration_files = list((ROOT / "supabase" / "migrations").glob("*.sql"))
        self.assertGreaterEqual(len(migration_files), 1)
        sql = "\n".join(path.read_text(encoding="utf-8").lower() for path in migration_files)

        for table in [
            "stock_selection_runs",
            "stock_selection_results",
            "stock_selection_prices",
            "stock_selection_performance",
            "stock_selection_job_runs",
        ]:
            self.assertIn(table, sql)
        self.assertIn("enable row level security", sql)
        self.assertIn("grant select", sql)
        self.assertIn("security_invoker", sql)
        self.assertIn("create or replace view public.dashboard_runs_index", sql)
        self.assertIn("create or replace view public.dashboard_runs", sql)
        self.assertIn("price_points", sql)
        self.assertIn("stock_selection_prices pr", sql)
        self.assertIn("create or replace view public.v_selection_runs_public", sql)
        self.assertIn("create or replace view public.v_selection_results_public", sql)
        self.assertIn("create or replace view public.v_selection_performance_public", sql)
        self.assertIn("create or replace view public.v_selection_summary_by_run_public", sql)
        self.assertIn("create or replace view public.v_selection_strategy_effectiveness_public", sql)
        self.assertIn("revoke all on table", sql)
        self.assertIn("service role can manage stock selection jobs", sql)
        self.assertIn("revoke all on table public.stock_selection_job_runs from anon", sql)


if __name__ == "__main__":
    unittest.main()
