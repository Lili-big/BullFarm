from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import shutil
import statistics
import sys
import time
from contextlib import contextmanager
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_CONFIG_PATH = PROJECT_ROOT / "config" / "validation_rules.json"

DEFAULT_CONFIG: dict[str, Any] = {
    "validation": {
        "output_excel": "data/validation/output/stock_selection_log.xlsx",
        "report_dir": "data/validation/output/reports",
        "backup_dir": "data/validation/output/backup",
        "default_offsets": [0, 1, 2, 3, 5, 10],
        "default_mode": "selection_price",
        "holding_days": 3,
        "stop_loss_pct": -3.0,
        "take_profit_pct": 5.0,
        "auto_backup_excel": True,
        "allow_force_overwrite": False,
    },
    "excel": {
        "freeze_header": True,
        "auto_adjust_column_width": True,
        "highlight_positive_return": True,
        "highlight_negative_return": True,
        "highlight_failed_result": True,
        "date_format": "yyyy-mm-dd",
        "percent_decimal_places": 2,
    },
}

SHEET_COLUMNS: dict[str, list[str]] = {
    "selection_runs": [
        "run_id",
        "selection_date",
        "selection_time",
        "strategy_version",
        "market_env",
        "total_selected_count",
        "data_source",
        "operator",
        "notes",
        "created_at",
    ],
    "selected_stocks": [
        "run_id",
        "selection_date",
        "selection_time",
        "stock_code",
        "stock_name",
        "sector",
        "industry",
        "concept",
        "strategy_label",
        "participation_level",
        "suggested_action",
        "selection_price",
        "selection_close",
        "selection_reason",
        "stop_loss_price",
        "take_profit_price",
        "risk_note",
        "total_score",
        "rank_in_run",
        "price_missing",
        "created_at",
    ],
    "future_prices": [
        "run_id",
        "selection_date",
        "stock_code",
        "stock_name",
        "trading_day_offset",
        "price_date",
        "open",
        "high",
        "low",
        "close",
        "volume",
        "amount",
        "turnover_rate",
        "is_suspended",
        "data_source",
        "updated_at",
    ],
    "performance": [
        "run_id",
        "selection_date",
        "stock_code",
        "stock_name",
        "sector",
        "industry",
        "concept",
        "strategy_label",
        "participation_level",
        "total_score",
        "rank_in_run",
        "selection_price",
        "latest_price",
        "latest_price_date",
        "return_latest_pct",
        "return_t1_close_pct",
        "return_t2_close_pct",
        "return_t3_close_pct",
        "return_t5_close_pct",
        "return_t10_close_pct",
        "max_high_3d",
        "max_gain_3d_pct",
        "min_low_3d",
        "max_drawdown_3d_pct",
        "max_high_5d",
        "max_gain_5d_pct",
        "min_low_5d",
        "max_drawdown_5d_pct",
        "hit_stop_loss",
        "hit_take_profit",
        "uncertain_intraday_order",
        "is_profitable_latest",
        "is_profitable_t1",
        "is_profitable_t2",
        "is_profitable_t3",
        "is_profitable_t5",
        "result_label",
        "failure_reason",
        "data_status",
        "updated_at",
    ],
    "summary_by_run": [
        "run_id",
        "selection_date",
        "strategy_version",
        "market_env",
        "total_selected_count",
        "valid_stock_count",
        "profitable_count_t1",
        "profitable_count_t2",
        "profitable_count_t3",
        "profitable_count_t5",
        "win_rate_t1",
        "win_rate_t2",
        "win_rate_t3",
        "win_rate_t5",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "median_return_t3_pct",
        "max_return_t3_pct",
        "min_return_t3_pct",
        "avg_max_drawdown_3d_pct",
        "hit_stop_loss_count",
        "hit_take_profit_count",
        "best_stock",
        "worst_stock",
        "conclusion",
        "updated_at",
    ],
    "summary_by_stock": [
        "stock_code",
        "stock_name",
        "selected_times",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "win_rate_t3",
        "avg_max_drawdown_3d_pct",
        "best_run_id",
        "worst_run_id",
        "last_selected_date",
        "conclusion",
    ],
    "summary_by_sector": [
        "sector",
        "selected_count",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "win_rate_t3",
        "avg_max_drawdown_3d_pct",
        "best_stock",
        "worst_stock",
        "conclusion",
    ],
    "summary_by_strategy_label": [
        "strategy_label",
        "selected_count",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "win_rate_t3",
        "avg_max_drawdown_3d_pct",
        "hit_stop_loss_rate",
        "hit_take_profit_rate",
        "conclusion",
    ],
    "summary_by_participation_level": [
        "participation_level",
        "selected_count",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "win_rate_t3",
        "avg_max_drawdown_3d_pct",
        "hit_stop_loss_rate",
        "hit_take_profit_rate",
        "conclusion",
    ],
    "comparison": [
        "group_by",
        "group_name",
        "selected_count",
        "valid_count",
        "win_rate_t1",
        "win_rate_t2",
        "win_rate_t3",
        "win_rate_t5",
        "avg_return_t1_pct",
        "avg_return_t2_pct",
        "avg_return_t3_pct",
        "avg_return_t5_pct",
        "median_return_t3_pct",
        "max_return_t3_pct",
        "min_return_t3_pct",
        "avg_max_drawdown_3d_pct",
        "hit_stop_loss_rate",
        "hit_take_profit_rate",
        "conclusion",
    ],
    "config_snapshot": ["key", "value", "updated_at"],
}


class ValidationError(RuntimeError):
    pass


@contextmanager
def workbook_lock(path: Path, timeout_seconds: int = 120) -> Any:
    lock_path = path.with_suffix(path.suffix + ".lock")
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    start = time.monotonic()
    fd: int | None = None
    while True:
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.write(fd, f"{os.getpid()} {now_text()}".encode("utf-8"))
            break
        except FileExistsError:
            if time.monotonic() - start > timeout_seconds:
                raise ValidationError(f"Validation workbook is locked: {lock_path}")
            time.sleep(0.25)
    try:
        yield
    finally:
        if fd is not None:
            os.close(fd)
        try:
            lock_path.unlink()
        except FileNotFoundError:
            pass


@dataclass
class PriceRecord:
    trade_date: str
    stock_code: str
    open: float | None
    high: float | None
    low: float | None
    close: float | None
    volume: float | None = None
    amount: float | None = None
    turnover_rate: float | None = None
    is_suspended: bool = False


def load_config(path: Path | None = None) -> dict[str, Any]:
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if path and path.exists():
        override = json.loads(path.read_text(encoding="utf-8-sig"))
        deep_update(config, override)
    return config


def deep_update(base: dict[str, Any], override: dict[str, Any]) -> None:
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_update(base[key], value)
        else:
            base[key] = value


def resolve_project_path(value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else PROJECT_ROOT / path


def resolve_related_path(value: str | Path, workbook_path: Path | None = None) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    if workbook_path is not None:
        try:
            workbook_path.resolve().relative_to(PROJECT_ROOT.resolve())
        except ValueError:
            return workbook_path.parent / path
    return PROJECT_ROOT / path


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_date_text(value: str | None) -> str:
    if not value:
        return date.today().strftime("%Y-%m-%d")
    cleaned = str(value).strip()
    if re.fullmatch(r"\d{8}", cleaned):
        return f"{cleaned[:4]}-{cleaned[4:6]}-{cleaned[6:]}"
    return datetime.strptime(cleaned, "%Y-%m-%d").strftime("%Y-%m-%d")


def infer_selection_date(candidates_path: Path, explicit: str | None) -> str:
    if explicit:
        return parse_date_text(explicit)
    match = re.search(r"(\d{8})", candidates_path.name)
    if match:
        return parse_date_text(match.group(1))
    return date.today().strftime("%Y-%m-%d")


def normalize_strategy_version(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_]+", "_", value.strip() or "v1_0")


def generate_run_id(selection_date: str | None = None, strategy_version: str = "v1_0") -> str:
    day = parse_date_text(selection_date).replace("-", "") if selection_date else date.today().strftime("%Y%m%d")
    clock = datetime.now().strftime("%H%M%S")
    return f"{day}_{clock}_{normalize_strategy_version(strategy_version)}"


def normalize_stock_code(value: Any) -> str:
    raw = str(value or "").strip().upper()
    if not raw:
        return ""
    raw = raw.replace(".XSHG", ".SH").replace(".XSHE", ".SZ")
    if re.fullmatch(r"\d{6}\.(SH|SZ|BJ)", raw):
        return raw
    digits = re.sub(r"\D", "", raw)
    if len(digits) < 6:
        digits = digits.zfill(6)
    elif len(digits) > 6:
        digits = digits[-6:]
    if digits.startswith(("60", "68", "90")):
        suffix = "SH"
    elif digits.startswith(("00", "30", "20")):
        suffix = "SZ"
    elif digits.startswith(("43", "83", "87", "88", "92")):
        suffix = "BJ"
    else:
        suffix = "SZ"
    return f"{digits}.{suffix}"


def plain_stock_code(value: str) -> str:
    return normalize_stock_code(value).split(".")[0]


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    cleaned = str(value).replace(",", "").replace("%", "").strip()
    if cleaned in {"", "-", "--", "None", "nan", "NaN"}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def as_int(value: Any, default: int = 0) -> int:
    number = as_float(value)
    return int(number) if number is not None else default


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "是", "停牌"}


def pct_change(future: float | None, base: float | None) -> float | None:
    if future is None or base is None or base == 0:
        return None
    return round((future - base) / base * 100, 4)


def average(values: Iterable[float | None]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    if not cleaned:
        return None
    return round(sum(cleaned) / len(cleaned), 4)


def median(values: Iterable[float | None]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    if not cleaned:
        return None
    return round(statistics.median(cleaned), 4)


def win_rate(values: Iterable[float | None]) -> float | None:
    cleaned = [float(v) for v in values if v is not None]
    if not cleaned:
        return None
    return round(sum(1 for v in cleaned if v > 0) / len(cleaned) * 100, 4)


def bool_rate(values: Iterable[Any]) -> float | None:
    cleaned = list(values)
    if not cleaned:
        return None
    return round(sum(1 for value in cleaned if as_bool(value)) / len(cleaned) * 100, 4)


def parse_offsets(value: str | None, config: dict[str, Any]) -> list[int]:
    if not value:
        return [int(v) for v in config["validation"]["default_offsets"]]
    offsets: list[int] = []
    for part in value.split(","):
        text = part.strip().upper().lstrip("T")
        if text:
            offsets.append(int(text))
    return sorted(set(offsets))


def offset_key(offset: int | str) -> str:
    text = str(offset).strip().upper()
    return text if text.startswith("T") else f"T{text}"


def offset_number(value: Any) -> int | None:
    text = str(value or "").strip().upper().lstrip("T")
    return int(text) if text.isdigit() else None


def read_csv_records(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv_records(path: Path, records: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        default = workbook.active
        default.title = "selection_runs"
    for sheet_name, columns in SHEET_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        ws = workbook[sheet_name]
        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            for idx, column in enumerate(columns, start=1):
                ws.cell(row=1, column=idx, value=column)
        else:
            current = [cell.value for cell in ws[1]]
            if current != columns:
                for idx, column in enumerate(columns, start=1):
                    ws.cell(row=1, column=idx, value=column)
    for sheet in list(workbook.sheetnames):
        if sheet not in SHEET_COLUMNS:
            del workbook[sheet]
    return workbook


def worksheet_records(workbook: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    ws = workbook[sheet_name]
    headers = [str(cell.value) for cell in ws[1]]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or value == "" for value in row):
            continue
        records.append({headers[idx]: row[idx] if idx < len(row) else "" for idx in range(len(headers))})
    return records


def set_records(workbook: Workbook, sheet_name: str, records: list[dict[str, Any]]) -> None:
    ws = workbook[sheet_name]
    ws.delete_rows(1, ws.max_row)
    columns = SHEET_COLUMNS[sheet_name]
    for idx, column in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=column)
    for record in records:
        ws.append([record.get(column, "") for column in columns])


def append_records(workbook: Workbook, sheet_name: str, records: list[dict[str, Any]]) -> None:
    ws = workbook[sheet_name]
    columns = SHEET_COLUMNS[sheet_name]
    for record in records:
        ws.append([record.get(column, "") for column in columns])


def backup_workbook(path: Path, config: dict[str, Any]) -> Path | None:
    if not path.exists() or not config["validation"].get("auto_backup_excel", True):
        return None
    backup_dir = resolve_related_path(config["validation"]["backup_dir"], path)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def save_workbook(workbook: Workbook, path: Path, config: dict[str, Any], backup: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if backup:
        backup_workbook(path, config)
    apply_excel_style(workbook, config)
    workbook.save(path)


def apply_excel_style(workbook: Workbook, config: dict[str, Any]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    failed_fill = PatternFill("solid", fgColor="F4CCCC")
    data_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet_name, columns in SHEET_COLUMNS.items():
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2" if config["excel"].get("freeze_header", True) else None
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        if config["excel"].get("auto_adjust_column_width", True):
            for idx, column in enumerate(columns, start=1):
                max_len = len(column)
                for cell in ws[get_column_letter(idx)]:
                    max_len = max(max_len, len(str(cell.value or "")))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 42)
        for idx, column in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            if column.endswith("_pct") or column.startswith("win_rate") or column.endswith("_rate"):
                for cell in ws[letter][1:]:
                    cell.number_format = "0.00"
                if config["excel"].get("highlight_positive_return", True):
                    ws.conditional_formatting.add(
                        f"{letter}2:{letter}{max(ws.max_row, 2)}",
                        CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="008000")),
                    )
                if config["excel"].get("highlight_negative_return", True):
                    ws.conditional_formatting.add(
                        f"{letter}2:{letter}{max(ws.max_row, 2)}",
                        CellIsRule(operator="lessThan", formula=["0"], font=Font(color="C00000")),
                    )
            elif column in {"selection_price", "selection_close", "latest_price", "open", "high", "low", "close"}:
                for cell in ws[letter][1:]:
                    cell.number_format = "0.00"
            elif column.endswith("_date") or column in {"created_at", "updated_at", "selection_time"}:
                for cell in ws[letter][1:]:
                    cell.number_format = "yyyy-mm-dd"
        if sheet_name == "performance" and ws.max_row >= 2:
            label_idx = columns.index("result_label") + 1
            for row in range(2, ws.max_row + 1):
                label = ws.cell(row=row, column=label_idx).value
                if label == "失败":
                    for cell in ws[row]:
                        cell.fill = failed_fill
                elif label in {"数据不足", "停牌无法判断"}:
                    for cell in ws[row]:
                        cell.fill = data_fill


def remove_run_records(workbook: Workbook, run_id: str, sheets: Iterable[str]) -> None:
    for sheet in sheets:
        records = worksheet_records(workbook, sheet)
        if "run_id" not in SHEET_COLUMNS[sheet]:
            continue
        set_records(workbook, sheet, [record for record in records if str(record.get("run_id")) != run_id])


def config_snapshot_records(config: dict[str, Any]) -> list[dict[str, Any]]:
    stamp = now_text()
    rows: list[dict[str, Any]] = []
    for section, values in config.items():
        if isinstance(values, dict):
            for key, value in values.items():
                rows.append({"key": f"{section}.{key}", "value": json.dumps(value, ensure_ascii=False), "updated_at": stamp})
        else:
            rows.append({"key": section, "value": json.dumps(values, ensure_ascii=False), "updated_at": stamp})
    return rows


def candidate_lookup(candidates: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in candidates:
        code = normalize_stock_code(row.get("symbol") or row.get("stock_code") or row.get("代码"))
        if code:
            result[code] = row
    return result


def build_snapshot_rows(
    scores: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    run_id: str,
    selection_date: str,
    selection_time: str,
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    lookup = candidate_lookup(candidates)
    stop_loss_pct = as_float(config["validation"]["stop_loss_pct"]) or -3.0
    take_profit_pct = as_float(config["validation"]["take_profit_pct"]) or 5.0
    rows: list[dict[str, Any]] = []
    for fallback_rank, score in enumerate(scores, start=1):
        code = normalize_stock_code(score.get("symbol") or score.get("stock_code"))
        candidate = lookup.get(code, {})
        selection_close = as_float(candidate.get("close") or candidate.get("selection_close"))
        selection_price = selection_close
        price_missing = selection_price is None
        stop_loss = round(selection_price * (1 + stop_loss_pct / 100), 4) if selection_price is not None else ""
        take_profit = round(selection_price * (1 + take_profit_pct / 100), 4) if selection_price is not None else ""
        rows.append(
            {
                "run_id": run_id,
                "selection_date": selection_date,
                "selection_time": selection_time,
                "stock_code": code,
                "stock_name": score.get("name") or candidate.get("name") or "",
                "sector": score.get("sector") or candidate.get("sector") or "",
                "industry": candidate.get("industry") or candidate.get("sector") or score.get("sector") or "",
                "concept": candidate.get("concept") or "",
                "strategy_label": score.get("buy_model") or "",
                "participation_level": score.get("decision") or "",
                "suggested_action": score.get("decision") or "",
                "selection_price": selection_price if selection_price is not None else "",
                "selection_close": selection_close if selection_close is not None else "",
                "selection_reason": score.get("notes") or score.get("plan") or "",
                "stop_loss_price": stop_loss,
                "take_profit_price": take_profit,
                "risk_note": score.get("risks") or score.get("hard_rejects") or "",
                "total_score": as_int(score.get("total_score")),
                "rank_in_run": as_int(score.get("rank"), fallback_rank),
                "price_missing": price_missing,
                "created_at": now_text(),
            }
        )
    return rows


def command_snapshot(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        selection_date = infer_selection_date(args.candidates, args.selection_date)
        strategy_version = normalize_strategy_version(args.strategy_version)
        run_id = args.run_id or generate_run_id(selection_date, strategy_version)
        existing_run_ids = {str(row.get("run_id")) for row in worksheet_records(workbook, "selection_runs")}
        if run_id in existing_run_ids:
            if not args.force:
                raise ValidationError(f"run_id already exists: {run_id}. Use --force to replace this run.")
            remove_run_records(
                workbook,
                run_id,
                [
                    "selection_runs",
                    "selected_stocks",
                    "future_prices",
                    "performance",
                    "summary_by_run",
                ],
            )
        scores = read_csv_records(args.scores)
        candidates = read_csv_records(args.candidates)
        selection_time = datetime.now().strftime("%H:%M:%S")
        stock_rows = build_snapshot_rows(scores, candidates, run_id, selection_date, selection_time, config)
        run_row = {
            "run_id": run_id,
            "selection_date": selection_date,
            "selection_time": selection_time,
            "strategy_version": strategy_version,
            "market_env": args.market_env,
            "total_selected_count": len(stock_rows),
            "data_source": f"scores={args.scores}; candidates={args.candidates}",
            "operator": args.operator,
            "notes": args.notes,
            "created_at": now_text(),
        }
        append_records(workbook, "selection_runs", [run_row])
        append_records(workbook, "selected_stocks", stock_rows)
        set_records(workbook, "config_snapshot", config_snapshot_records(config))
        save_workbook(workbook, output_excel, config)
    print(f"Validation snapshot saved: {output_excel}")
    print(f"run_id: {run_id}")
    print(f"selected stocks: {len(stock_rows)}")
    return 0


def read_price_file(path: Path) -> dict[str, list[PriceRecord]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True)
        ws = workbook.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
    else:
        rows = read_csv_records(path)
    by_code: dict[str, list[PriceRecord]] = defaultdict(list)
    for row in rows:
        code = normalize_stock_code(row.get("stock_code") or row.get("symbol") or row.get("代码"))
        if not code:
            continue
        trade_date = row.get("trade_date") or row.get("date") or row.get("日期")
        trade_date_text = parse_date_text(str(trade_date)[:10].replace("/", "-"))
        record = PriceRecord(
            trade_date=trade_date_text,
            stock_code=code,
            open=as_float(row.get("open") or row.get("开盘")),
            high=as_float(row.get("high") or row.get("最高")),
            low=as_float(row.get("low") or row.get("最低")),
            close=as_float(row.get("close") or row.get("收盘")),
            volume=as_float(row.get("volume") or row.get("成交量")),
            amount=as_float(row.get("amount") or row.get("成交额")),
            turnover_rate=as_float(row.get("turnover_rate") or row.get("换手率")),
            is_suspended=as_bool(row.get("is_suspended")),
        )
        by_code[code].append(record)
    for records in by_code.values():
        records.sort(key=lambda item: item.trade_date)
    return by_code


def fetch_akshare_prices(stock_code: str, selection_date: str, end_date: str | None = None) -> list[PriceRecord]:
    try:
        import akshare as ak  # type: ignore
    except ImportError as exc:
        raise ValidationError("akshare is required for live price updates. Use --price-file for offline updates.") from exc
    end = end_date or (date.today() + timedelta(days=30)).strftime("%Y%m%d")
    start = selection_date.replace("-", "")
    raw = ak.stock_zh_a_hist(symbol=plain_stock_code(stock_code), period="daily", start_date=start, end_date=end, adjust="qfq")
    records: list[PriceRecord] = []
    for _, row in raw.iterrows():
        trade_date = parse_date_text(str(row.get("日期")))
        records.append(
            PriceRecord(
                trade_date=trade_date,
                stock_code=normalize_stock_code(stock_code),
                open=as_float(row.get("开盘")),
                high=as_float(row.get("最高")),
                low=as_float(row.get("最低")),
                close=as_float(row.get("收盘")),
                volume=as_float(row.get("成交量")),
                amount=as_float(row.get("成交额")),
                turnover_rate=as_float(row.get("换手率")),
                is_suspended=False,
            )
        )
    records.sort(key=lambda item: item.trade_date)
    return records


def records_for_offsets(records: list[PriceRecord], selection_date: str, offsets: list[int]) -> dict[int, PriceRecord]:
    eligible = [record for record in records if record.trade_date >= selection_date]
    eligible.sort(key=lambda item: item.trade_date)
    return {idx: record for idx, record in enumerate(eligible) if idx in set(offsets)}


def command_update_prices(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        offsets = parse_offsets(args.offsets, config)
        run_ids = selected_run_ids(workbook, args.run_id, args.all)
        selected = [row for row in worksheet_records(workbook, "selected_stocks") if row.get("run_id") in run_ids]
        if not selected:
            raise ValidationError("No selected stocks found for price update.")
        price_file_records = read_price_file(args.price_file) if args.price_file else None
        existing = [row for row in worksheet_records(workbook, "future_prices") if row.get("run_id") not in run_ids]
        updated_rows: list[dict[str, Any]] = []
        stamp = now_text()
        for selected_row in selected:
            code = normalize_stock_code(selected_row.get("stock_code"))
            selection_date = parse_date_text(str(selected_row.get("selection_date")))
            if price_file_records is not None:
                price_records = price_file_records.get(code, [])
                source = f"file:{args.price_file}"
            else:
                price_records = fetch_akshare_prices(code, selection_date, args.end_date)
                source = "akshare.stock_zh_a_hist"
            by_offset = records_for_offsets(price_records, selection_date, offsets)
            for offset in offsets:
                record = by_offset.get(offset)
                if not record:
                    continue
                updated_rows.append(
                    {
                        "run_id": selected_row.get("run_id"),
                        "selection_date": selection_date,
                        "stock_code": code,
                        "stock_name": selected_row.get("stock_name"),
                        "trading_day_offset": offset_key(offset),
                        "price_date": record.trade_date,
                        "open": record.open if record.open is not None else "",
                        "high": record.high if record.high is not None else "",
                        "low": record.low if record.low is not None else "",
                        "close": record.close if record.close is not None else "",
                        "volume": record.volume if record.volume is not None else "",
                        "amount": record.amount if record.amount is not None else "",
                        "turnover_rate": record.turnover_rate if record.turnover_rate is not None else "",
                        "is_suspended": record.is_suspended,
                        "data_source": source,
                        "updated_at": stamp,
                    }
                )
        set_records(workbook, "future_prices", existing + updated_rows)
        save_workbook(workbook, output_excel, config)
    print(f"Updated future prices: {len(updated_rows)} rows")
    return 0


def selected_run_ids(workbook: Workbook, run_id: str | None, all_runs: bool = False) -> set[str]:
    runs = [str(row.get("run_id")) for row in worksheet_records(workbook, "selection_runs") if row.get("run_id")]
    if run_id:
        if run_id not in runs:
            raise ValidationError(f"Unknown run_id: {run_id}")
        return {run_id}
    if all_runs or runs:
        return set(runs)
    raise ValidationError("No run_id provided and no runs exist.")


def price_index(future_prices: list[dict[str, Any]]) -> dict[tuple[str, str], dict[int, dict[str, Any]]]:
    index: dict[tuple[str, str], dict[int, dict[str, Any]]] = defaultdict(dict)
    for row in future_prices:
        offset = offset_number(row.get("trading_day_offset"))
        if offset is None:
            continue
        key = (str(row.get("run_id")), normalize_stock_code(row.get("stock_code")))
        index[key][offset] = row
    return index


def calculate_performance_row(
    selected: dict[str, Any],
    prices_by_offset: dict[int, dict[str, Any]],
    config: dict[str, Any],
) -> dict[str, Any]:
    base = as_float(selected.get("selection_price") or selected.get("selection_close"))
    stock_code = normalize_stock_code(selected.get("stock_code"))
    suspended = any(as_bool(row.get("is_suspended")) for row in prices_by_offset.values())
    price_missing = as_bool(selected.get("price_missing")) or base is None
    close_by_offset = {offset: as_float(row.get("close")) for offset, row in prices_by_offset.items()}
    high_by_offset = {offset: as_float(row.get("high")) for offset, row in prices_by_offset.items()}
    low_by_offset = {offset: as_float(row.get("low")) for offset, row in prices_by_offset.items()}
    latest_offset = max((offset for offset, value in close_by_offset.items() if value is not None), default=None)
    latest_row = prices_by_offset.get(latest_offset, {}) if latest_offset is not None else {}
    latest_price = as_float(latest_row.get("close"))
    returns = {offset: pct_change(close_by_offset.get(offset), base) for offset in [1, 2, 3, 5, 10]}
    highs3 = [high_by_offset.get(offset) for offset in [1, 2, 3] if high_by_offset.get(offset) is not None]
    lows3 = [low_by_offset.get(offset) for offset in [1, 2, 3] if low_by_offset.get(offset) is not None]
    highs5 = [high_by_offset.get(offset) for offset in [1, 2, 3, 4, 5] if high_by_offset.get(offset) is not None]
    lows5 = [low_by_offset.get(offset) for offset in [1, 2, 3, 4, 5] if low_by_offset.get(offset) is not None]
    max_high_3d = max(highs3) if highs3 else None
    min_low_3d = min(lows3) if lows3 else None
    max_high_5d = max(highs5) if highs5 else None
    min_low_5d = min(lows5) if lows5 else None
    holding_days = as_int(config["validation"]["holding_days"], 3)
    stop_loss_price = as_float(selected.get("stop_loss_price"))
    take_profit_price = as_float(selected.get("take_profit_price"))
    if base is not None:
        if stop_loss_price is None:
            stop_loss_price = base * (1 + (as_float(config["validation"]["stop_loss_pct"]) or -3.0) / 100)
        if take_profit_price is None:
            take_profit_price = base * (1 + (as_float(config["validation"]["take_profit_pct"]) or 5.0) / 100)
    holding_offsets = range(1, holding_days + 1)
    hit_stop = any(
        low_by_offset.get(offset) is not None and stop_loss_price is not None and low_by_offset[offset] <= stop_loss_price
        for offset in holding_offsets
    )
    hit_take = any(
        high_by_offset.get(offset) is not None and take_profit_price is not None and high_by_offset[offset] >= take_profit_price
        for offset in holding_offsets
    )
    uncertain = hit_stop and hit_take
    data_status = "complete"
    failure_reason = ""
    if price_missing:
        data_status = "price_missing"
        result_label = "数据不足"
        failure_reason = "selection_price missing"
    elif suspended:
        data_status = "suspended"
        result_label = "停牌无法判断"
        failure_reason = "suspended"
    elif returns[3] is None:
        data_status = "insufficient_future_prices"
        result_label = "数据不足"
        failure_reason = "T3 price missing"
    else:
        expected = {0, 1, 2, 3, 5, 10}
        present = {offset for offset, value in close_by_offset.items() if value is not None}
        data_status = "complete" if expected.issubset(present) else "partial"
        result_label = classify_result_label(returns[3])
    return {
        "run_id": selected.get("run_id"),
        "selection_date": selected.get("selection_date"),
        "stock_code": stock_code,
        "stock_name": selected.get("stock_name"),
        "sector": selected.get("sector"),
        "industry": selected.get("industry"),
        "concept": selected.get("concept"),
        "strategy_label": selected.get("strategy_label"),
        "participation_level": selected.get("participation_level"),
        "total_score": as_int(selected.get("total_score")),
        "rank_in_run": as_int(selected.get("rank_in_run")),
        "selection_price": base if base is not None else "",
        "latest_price": latest_price if latest_price is not None else "",
        "latest_price_date": latest_row.get("price_date", ""),
        "return_latest_pct": pct_change(latest_price, base),
        "return_t1_close_pct": returns[1],
        "return_t2_close_pct": returns[2],
        "return_t3_close_pct": returns[3],
        "return_t5_close_pct": returns[5],
        "return_t10_close_pct": returns[10],
        "max_high_3d": max_high_3d if max_high_3d is not None else "",
        "max_gain_3d_pct": pct_change(max_high_3d, base),
        "min_low_3d": min_low_3d if min_low_3d is not None else "",
        "max_drawdown_3d_pct": pct_change(min_low_3d, base),
        "max_high_5d": max_high_5d if max_high_5d is not None else "",
        "max_gain_5d_pct": pct_change(max_high_5d, base),
        "min_low_5d": min_low_5d if min_low_5d is not None else "",
        "max_drawdown_5d_pct": pct_change(min_low_5d, base),
        "hit_stop_loss": hit_stop,
        "hit_take_profit": hit_take,
        "uncertain_intraday_order": uncertain,
        "is_profitable_latest": latest_price is not None and base is not None and latest_price > base,
        "is_profitable_t1": returns[1] is not None and returns[1] > 0,
        "is_profitable_t2": returns[2] is not None and returns[2] > 0,
        "is_profitable_t3": returns[3] is not None and returns[3] > 0,
        "is_profitable_t5": returns[5] is not None and returns[5] > 0,
        "result_label": result_label,
        "failure_reason": failure_reason,
        "data_status": data_status,
        "updated_at": now_text(),
    }


def classify_result_label(return_t3: float | None) -> str:
    if return_t3 is None:
        return "数据不足"
    if return_t3 > 3:
        return "成功"
    if return_t3 > 0:
        return "小赚"
    if return_t3 >= -1:
        return "基本持平"
    if return_t3 >= -3:
        return "小亏"
    return "失败"


def command_analyze(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        run_ids = selected_run_ids(workbook, args.run_id, args.all)
        selected = [row for row in worksheet_records(workbook, "selected_stocks") if row.get("run_id") in run_ids]
        price_rows = worksheet_records(workbook, "future_prices")
        index = price_index(price_rows)
        new_perf = [
            calculate_performance_row(row, index.get((str(row.get("run_id")), normalize_stock_code(row.get("stock_code"))), {}), config)
            for row in selected
        ]
        existing_perf = [row for row in worksheet_records(workbook, "performance") if row.get("run_id") not in run_ids]
        all_perf = existing_perf + new_perf
        set_records(workbook, "performance", all_perf)
        refresh_summaries(workbook, all_perf)
        save_workbook(workbook, output_excel, config)
    print(f"Analyzed performance rows: {len(new_perf)}")
    return 0


def refresh_summaries(workbook: Workbook, performance_rows: list[dict[str, Any]]) -> None:
    runs = {str(row.get("run_id")): row for row in worksheet_records(workbook, "selection_runs")}
    set_records(workbook, "summary_by_run", build_summary_by_run(performance_rows, runs))
    set_records(workbook, "summary_by_stock", build_summary_by_stock(performance_rows))
    set_records(workbook, "summary_by_sector", build_dimension_summary(performance_rows, "sector", "summary_by_sector"))
    set_records(workbook, "summary_by_strategy_label", build_dimension_summary(performance_rows, "strategy_label", "summary_by_strategy_label"))
    set_records(
        workbook,
        "summary_by_participation_level",
        build_dimension_summary(performance_rows, "participation_level", "summary_by_participation_level"),
    )


def build_summary_by_run(rows: list[dict[str, Any]], runs: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    grouped = group_rows(rows, "run_id")
    results: list[dict[str, Any]] = []
    for run_id, items in sorted(grouped.items()):
        run = runs.get(run_id, {})
        valid_t3 = [as_float(row.get("return_t3_close_pct")) for row in items if as_float(row.get("return_t3_close_pct")) is not None]
        best = max(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else -999999)
        worst = min(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else 999999)
        results.append(
            {
                "run_id": run_id,
                "selection_date": run.get("selection_date") or first(items, "selection_date"),
                "strategy_version": run.get("strategy_version", ""),
                "market_env": run.get("market_env", ""),
                "total_selected_count": run.get("total_selected_count") or len(items),
                "valid_stock_count": len(valid_t3),
                "profitable_count_t1": count_positive(items, "return_t1_close_pct"),
                "profitable_count_t2": count_positive(items, "return_t2_close_pct"),
                "profitable_count_t3": count_positive(items, "return_t3_close_pct"),
                "profitable_count_t5": count_positive(items, "return_t5_close_pct"),
                "win_rate_t1": win_rate(row.get("return_t1_close_pct") for row in items),
                "win_rate_t2": win_rate(row.get("return_t2_close_pct") for row in items),
                "win_rate_t3": win_rate(row.get("return_t3_close_pct") for row in items),
                "win_rate_t5": win_rate(row.get("return_t5_close_pct") for row in items),
                "avg_return_t1_pct": average(row.get("return_t1_close_pct") for row in items),
                "avg_return_t2_pct": average(row.get("return_t2_close_pct") for row in items),
                "avg_return_t3_pct": average(row.get("return_t3_close_pct") for row in items),
                "avg_return_t5_pct": average(row.get("return_t5_close_pct") for row in items),
                "median_return_t3_pct": median(row.get("return_t3_close_pct") for row in items),
                "max_return_t3_pct": max(valid_t3) if valid_t3 else None,
                "min_return_t3_pct": min(valid_t3) if valid_t3 else None,
                "avg_max_drawdown_3d_pct": average(row.get("max_drawdown_3d_pct") for row in items),
                "hit_stop_loss_count": sum(1 for row in items if as_bool(row.get("hit_stop_loss"))),
                "hit_take_profit_count": sum(1 for row in items if as_bool(row.get("hit_take_profit"))),
                "best_stock": best.get("stock_name"),
                "worst_stock": worst.get("stock_name"),
                "conclusion": conclusion_for_group(items),
                "updated_at": now_text(),
            }
        )
    return results


def build_summary_by_stock(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped = group_rows(rows, "stock_code")
    results: list[dict[str, Any]] = []
    for code, items in sorted(grouped.items()):
        best = max(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else -999999)
        worst = min(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else 999999)
        results.append(
            {
                "stock_code": code,
                "stock_name": first(items, "stock_name"),
                "selected_times": len(items),
                "avg_return_t1_pct": average(row.get("return_t1_close_pct") for row in items),
                "avg_return_t2_pct": average(row.get("return_t2_close_pct") for row in items),
                "avg_return_t3_pct": average(row.get("return_t3_close_pct") for row in items),
                "avg_return_t5_pct": average(row.get("return_t5_close_pct") for row in items),
                "win_rate_t3": win_rate(row.get("return_t3_close_pct") for row in items),
                "avg_max_drawdown_3d_pct": average(row.get("max_drawdown_3d_pct") for row in items),
                "best_run_id": best.get("run_id"),
                "worst_run_id": worst.get("run_id"),
                "last_selected_date": max(str(row.get("selection_date")) for row in items),
                "conclusion": conclusion_for_group(items),
            }
        )
    return results


def build_dimension_summary(rows: list[dict[str, Any]], key: str, sheet_name: str) -> list[dict[str, Any]]:
    grouped = group_rows(rows, key)
    label = SHEET_COLUMNS[sheet_name][0]
    results: list[dict[str, Any]] = []
    for name, items in sorted(grouped.items()):
        best = max(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else -999999)
        worst = min(items, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else 999999)
        record = {
            label: name or "未填写",
            "selected_count": len(items),
            "avg_return_t1_pct": average(row.get("return_t1_close_pct") for row in items),
            "avg_return_t2_pct": average(row.get("return_t2_close_pct") for row in items),
            "avg_return_t3_pct": average(row.get("return_t3_close_pct") for row in items),
            "avg_return_t5_pct": average(row.get("return_t5_close_pct") for row in items),
            "win_rate_t3": win_rate(row.get("return_t3_close_pct") for row in items),
            "avg_max_drawdown_3d_pct": average(row.get("max_drawdown_3d_pct") for row in items),
            "best_stock": best.get("stock_name"),
            "worst_stock": worst.get("stock_name"),
            "hit_stop_loss_rate": bool_rate(row.get("hit_stop_loss") for row in items),
            "hit_take_profit_rate": bool_rate(row.get("hit_take_profit") for row in items),
            "conclusion": conclusion_for_group(items),
        }
        results.append({column: record.get(column, "") for column in SHEET_COLUMNS[sheet_name]})
    return results


def group_rows(rows: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get(key) or "未填写")].append(row)
    return grouped


def first(rows: list[dict[str, Any]], key: str) -> Any:
    for row in rows:
        value = row.get(key)
        if value not in {None, ""}:
            return value
    return ""


def count_positive(rows: list[dict[str, Any]], key: str) -> int:
    return sum(1 for row in rows if (as_float(row.get(key)) or 0) > 0)


def conclusion_for_group(rows: list[dict[str, Any]]) -> str:
    t3_win_rate = win_rate(row.get("return_t3_close_pct") for row in rows)
    avg_t3 = average(row.get("return_t3_close_pct") for row in rows)
    if t3_win_rate is None or avg_t3 is None:
        return "数据不足，暂不判断。"
    if t3_win_rate >= 60 and avg_t3 > 0:
        return f"表现较好，T3 胜率 {t3_win_rate:.2f}%，平均收益为正。"
    if avg_t3 > 0:
        return "表现一般，平均收益为正但胜率仍需观察。"
    return "表现偏弱，T3 平均收益为负，需要复盘规则。"


def command_compare(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        performance_rows = filter_by_date(worksheet_records(workbook, "performance"), args.start_date, args.end_date)
        run_map = {str(row.get("run_id")): row for row in worksheet_records(workbook, "selection_runs")}
        group_by = args.group_by
        if group_by == "total_score_bucket":
            for row in performance_rows:
                row["total_score_bucket"] = score_bucket(row.get("total_score"))
        elif group_by == "market_env":
            for row in performance_rows:
                row["market_env"] = run_map.get(str(row.get("run_id")), {}).get("market_env", "未填写")
        if group_by not in {"run_id", "selection_date", "sector", "industry", "concept", "strategy_label", "participation_level", "market_env", "result_label", "total_score_bucket"}:
            raise ValidationError(f"Unsupported group-by: {group_by}")
        rows = build_comparison_rows(performance_rows, group_by)
        set_records(workbook, "comparison", rows)
        save_workbook(workbook, output_excel, config)
    print(f"Comparison rows: {len(rows)}")
    return 0


def build_comparison_rows(rows: list[dict[str, Any]], group_by: str) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for name, items in sorted(group_rows(rows, group_by).items()):
        valid_count = sum(1 for row in items if as_float(row.get("return_t3_close_pct")) is not None)
        results.append(
            {
                "group_by": group_by,
                "group_name": name,
                "selected_count": len(items),
                "valid_count": valid_count,
                "win_rate_t1": win_rate(row.get("return_t1_close_pct") for row in items),
                "win_rate_t2": win_rate(row.get("return_t2_close_pct") for row in items),
                "win_rate_t3": win_rate(row.get("return_t3_close_pct") for row in items),
                "win_rate_t5": win_rate(row.get("return_t5_close_pct") for row in items),
                "avg_return_t1_pct": average(row.get("return_t1_close_pct") for row in items),
                "avg_return_t2_pct": average(row.get("return_t2_close_pct") for row in items),
                "avg_return_t3_pct": average(row.get("return_t3_close_pct") for row in items),
                "avg_return_t5_pct": average(row.get("return_t5_close_pct") for row in items),
                "median_return_t3_pct": median(row.get("return_t3_close_pct") for row in items),
                "max_return_t3_pct": max([as_float(row.get("return_t3_close_pct")) for row in items if as_float(row.get("return_t3_close_pct")) is not None], default=None),
                "min_return_t3_pct": min([as_float(row.get("return_t3_close_pct")) for row in items if as_float(row.get("return_t3_close_pct")) is not None], default=None),
                "avg_max_drawdown_3d_pct": average(row.get("max_drawdown_3d_pct") for row in items),
                "hit_stop_loss_rate": bool_rate(row.get("hit_stop_loss") for row in items),
                "hit_take_profit_rate": bool_rate(row.get("hit_take_profit") for row in items),
                "conclusion": conclusion_for_group(items),
            }
        )
    return results


def filter_by_date(rows: list[dict[str, Any]], start_date: str | None, end_date: str | None) -> list[dict[str, Any]]:
    start = parse_date_text(start_date) if start_date else None
    end = parse_date_text(end_date) if end_date else None
    result = []
    for row in rows:
        selection_date = str(row.get("selection_date") or "")
        if start and selection_date < start:
            continue
        if end and selection_date > end:
            continue
        result.append(row)
    return result


def score_bucket(value: Any) -> str:
    score = as_float(value)
    if score is None or score < 50:
        return "50以下"
    if score < 65:
        return "50-65"
    if score < 80:
        return "65-80"
    return "80以上"


def command_report(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        run_id = args.run_id
        runs = {str(row.get("run_id")): row for row in worksheet_records(workbook, "selection_runs")}
        if run_id not in runs:
            raise ValidationError(f"Unknown run_id: {run_id}")
        summary = next((row for row in worksheet_records(workbook, "summary_by_run") if row.get("run_id") == run_id), None)
        perf = [row for row in worksheet_records(workbook, "performance") if row.get("run_id") == run_id]
        if not summary or not perf:
            raise ValidationError("No analysis data found. Run analyze first.")
        report_dir = resolve_related_path(args.report_dir or config["validation"]["report_dir"], output_excel)
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path = report_dir / f"report_{run_id}.md"
        report_path.write_text(build_markdown_report(run_id, runs[run_id], summary, perf), encoding="utf-8")
    print(f"Validation report: {report_path}")
    return 0


def build_markdown_report(run_id: str, run: dict[str, Any], summary: dict[str, Any], perf: list[dict[str, Any]]) -> str:
    top = sorted(perf, key=lambda row: as_float(row.get("return_t3_close_pct")) if as_float(row.get("return_t3_close_pct")) is not None else -999999, reverse=True)
    bottom = list(reversed(top))
    drawdown = sorted(perf, key=lambda row: as_float(row.get("max_drawdown_3d_pct")) if as_float(row.get("max_drawdown_3d_pct")) is not None else 999999)
    lines = [
        "# 选股结果验证报告",
        "",
        "## 基本信息",
        "",
        f"- run_id：{run_id}",
        f"- selection_date：{run.get('selection_date', '')}",
        f"- strategy_version：{run.get('strategy_version', '')}",
        f"- market_env：{run.get('market_env', '')}",
        f"- total_selected_count：{run.get('total_selected_count', '')}",
        "",
        "## 整体表现",
        "",
        f"- T1 胜率：{fmt(summary.get('win_rate_t1'))}%",
        f"- T2 胜率：{fmt(summary.get('win_rate_t2'))}%",
        f"- T3 胜率：{fmt(summary.get('win_rate_t3'))}%",
        f"- T5 胜率：{fmt(summary.get('win_rate_t5'))}%",
        f"- T1 平均收益：{fmt(summary.get('avg_return_t1_pct'))}%",
        f"- T2 平均收益：{fmt(summary.get('avg_return_t2_pct'))}%",
        f"- T3 平均收益：{fmt(summary.get('avg_return_t3_pct'))}%",
        f"- T5 平均收益：{fmt(summary.get('avg_return_t5_pct'))}%",
        f"- 未来 3 日平均最大回撤：{fmt(summary.get('avg_max_drawdown_3d_pct'))}%",
        f"- 止损触发数量：{summary.get('hit_stop_loss_count', '')}",
        f"- 止盈触发数量：{summary.get('hit_take_profit_count', '')}",
        "",
        "## 个股表现",
        "",
        "收益最高前 5：",
        *stock_lines(top[:5]),
        "",
        "收益最低前 5：",
        *stock_lines(bottom[:5]),
        "",
        "最大回撤前 5：",
        *drawdown_lines(drawdown[:5]),
        "",
        "数据不足股票：",
        *stock_lines([row for row in perf if row.get("result_label") in {"数据不足", "停牌无法判断"}][:10]),
        "",
        "## 结论",
        "",
        str(summary.get("conclusion") or "数据不足，暂不判断。"),
        "",
        "## 后续优化建议",
        "",
        "- 持续观察不同 strategy_label 和 participation_level 的 T3 胜率与平均收益。",
        "- 如果 T1 表现好但 T3 回落明显，后续需要缩短持有验证周期。",
        "- 如果最大回撤偏大，后续需要优化止损规则。",
        "",
        "说明：本报告只用于规则化复盘，不构成投资建议。",
        "",
    ]
    return "\n".join(lines)


def fmt(value: Any) -> str:
    number = as_float(value)
    return "NA" if number is None else f"{number:.2f}"


def stock_lines(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["- 无"]
    return [f"- {row.get('stock_code')} {row.get('stock_name')}：T3 {fmt(row.get('return_t3_close_pct'))}%，结果 {row.get('result_label')}" for row in rows]


def drawdown_lines(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["- 无"]
    return [f"- {row.get('stock_code')} {row.get('stock_name')}：3日最大回撤 {fmt(row.get('max_drawdown_3d_pct'))}%" for row in rows]


def command_list_runs(args: argparse.Namespace) -> int:
    config = load_config(args.config)
    output_excel = resolve_project_path(args.output_excel or config["validation"]["output_excel"])
    with workbook_lock(output_excel):
        workbook = ensure_workbook(output_excel)
        for row in worksheet_records(workbook, "selection_runs"):
            print(
                f"{row.get('run_id')}\t{row.get('selection_date')}\t"
                f"{row.get('market_env')}\t{row.get('total_selected_count')}"
            )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Validate stock selection results with forward prices and Excel reports.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="Validation config path.")
    parser.add_argument("--output-excel", type=Path, default=None, help="Validation workbook path.")
    sub = parser.add_subparsers(dest="command", required=True)

    snapshot = sub.add_parser("snapshot", help="Append one scored selection run to the validation workbook.")
    snapshot.add_argument("--scores", type=Path, default=PROJECT_ROOT / "outputs" / "selection_scores.csv")
    snapshot.add_argument("--candidates", type=Path, required=True)
    snapshot.add_argument("--run-id", default=None)
    snapshot.add_argument("--strategy-version", default="v1_0")
    snapshot.add_argument("--selection-date", default=None)
    snapshot.add_argument("--market-env", default="未填写")
    snapshot.add_argument("--operator", default="")
    snapshot.add_argument("--notes", default="")
    snapshot.add_argument("--force", action="store_true")
    snapshot.set_defaults(func=command_snapshot)

    update_prices = sub.add_parser("update-prices", help="Update T+N prices for one or all runs.")
    update_prices.add_argument("--run-id", default=None)
    update_prices.add_argument("--all", action="store_true")
    update_prices.add_argument("--offsets", default=None)
    update_prices.add_argument("--price-file", type=Path, default=None)
    update_prices.add_argument("--end-date", default=None)
    update_prices.set_defaults(func=command_update_prices)

    analyze = sub.add_parser("analyze", help="Analyze performance for one or all runs.")
    analyze.add_argument("--run-id", default=None)
    analyze.add_argument("--all", action="store_true")
    analyze.set_defaults(func=command_analyze)

    compare = sub.add_parser("compare", help="Compare historical performance by one dimension.")
    compare.add_argument("--group-by", default="strategy_label")
    compare.add_argument("--start-date", default=None)
    compare.add_argument("--end-date", default=None)
    compare.set_defaults(func=command_compare)

    report = sub.add_parser("report", help="Generate a Markdown validation report for one run.")
    report.add_argument("--run-id", required=True)
    report.add_argument("--report-dir", type=Path, default=None)
    report.set_defaults(func=command_report)

    list_runs = sub.add_parser("list-runs", help="List stored selection runs.")
    list_runs.set_defaults(func=command_list_runs)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except ValidationError as exc:
        print(f"Validation error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())


