from __future__ import annotations

import argparse
import csv
import json
import re
import statistics
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_SCORES_DIR = PROJECT_ROOT / "outputs"
DEFAULT_VALIDATION_WORKBOOK = PROJECT_ROOT / "data" / "validation" / "output" / "stock_selection_log.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "dashboard"
SCHEMA_VERSION = 1
LATEST_OFFSET_RANK = 999999

RETURN_FIELDS = [
    "return_latest_pct",
    "return_t1_close_pct",
    "return_t2_close_pct",
    "return_t3_close_pct",
    "return_t5_close_pct",
    "return_t10_close_pct",
]


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def to_json_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def as_float(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result


def as_int(value: Any, default: int = 0) -> int:
    number = as_float(value)
    return default if number is None else int(number)


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return clean_text(value).lower() in {"1", "true", "yes", "y", "success", "成功"}


def rounded(value: Any, digits: int = 2) -> float | None:
    number = as_float(value)
    return None if number is None else round(number, digits)


def pct_change(close_price: Any, base_price: Any) -> float | None:
    close = as_float(close_price)
    base = as_float(base_price)
    if close is None or base in {None, 0}:
        return None
    return rounded((close - base) / base * 100)


def compact_number(value: Any) -> int | float | None:
    number = as_float(value)
    if number is None:
        return None
    return int(number) if number.is_integer() else round(number, 4)


def date_key_from_text(value: Any) -> str:
    text = clean_text(value)
    match = re.search(r"(20\d{2})[-_/]?(\d{2})[-_/]?(\d{2})", text)
    return "".join(match.groups()) if match else ""


def display_date(date_key: str) -> str:
    if re.fullmatch(r"20\d{6}", date_key):
        return f"{date_key[:4]}-{date_key[4:6]}-{date_key[6:8]}"
    return date_key


def normalize_stock_code(value: Any) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    match = re.search(r"(\d{6})", text)
    if not match:
        return text
    code = match.group(1)
    suffix_match = re.search(r"\.(SH|SZ|BJ)$", text)
    if suffix_match:
        suffix = suffix_match.group(1)
    elif code.startswith(("6", "9")):
        suffix = "SH"
    elif code.startswith(("4", "8")):
        suffix = "BJ"
    else:
        suffix = "SZ"
    return f"{code}.{suffix}"


def symbol_from_code(value: Any) -> str:
    code = normalize_stock_code(value)
    match = re.search(r"(\d{6})", code)
    return match.group(1) if match else code


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def read_workbook_records(path: Path) -> dict[str, list[dict[str, Any]]]:
    if not path.exists():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        records: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = [clean_text(cell) for cell in next(rows)]
            except StopIteration:
                records[sheet_name] = []
                continue
            sheet_records: list[dict[str, Any]] = []
            for row in rows:
                if all(cell in {None, ""} for cell in row):
                    continue
                sheet_records.append(
                    {
                        headers[index]: to_json_scalar(row[index]) if index < len(row) else ""
                        for index in range(len(headers))
                        if headers[index]
                    }
                )
            records[sheet_name] = sheet_records
        return records
    finally:
        workbook.close()


def group_by(rows: Iterable[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[clean_text(row.get(key))].append(row)
    return grouped


def offset_sort_key(value: Any) -> int:
    text = clean_text(value).upper()
    if text == "LATEST":
        return LATEST_OFFSET_RANK
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else LATEST_OFFSET_RANK - 1


def first_present(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean_text(row.get(key))
        if value:
            return value
    return ""


def score_file_date(path: Path) -> str:
    for value in (path.name, path.parent.name, str(path)):
        found = date_key_from_text(value)
        if found:
            return found
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y%m%d")


def extract_referenced_score_paths(runs: Iterable[dict[str, Any]], project_root: Path) -> set[Path]:
    paths: set[Path] = set()
    for run in runs:
        source = clean_text(run.get("data_source"))
        match = re.search(r"scores=([^;]+)", source)
        if not match:
            continue
        raw_path = match.group(1).strip().replace("\\", "/")
        path = Path(raw_path)
        if not path.is_absolute():
            path = project_root / path
        paths.add(path.resolve())
    return paths


def discover_score_files(scores_dir: Path, referenced_paths: set[Path]) -> list[Path]:
    candidates = list(scores_dir.glob("selection_scores*.csv"))
    candidates.extend(path for path in scores_dir.glob("*/selection_scores*.csv") if path.parent.name != "latest")
    candidates.extend(path for path in scores_dir.glob("daily/*/selection_scores*.csv") if path.parent.name != "latest")
    files = sorted({path.resolve() for path in candidates})
    dated_keys = {score_file_date(path) for path in files if date_key_from_text(path.name)}
    result: list[Path] = []
    for path in files:
        if path in referenced_paths:
            continue
        if not date_key_from_text(path.name) and score_file_date(path) in dated_keys:
            continue
        result.append(path)
    return result


def build_price_points(price_rows: Iterable[dict[str, Any]], base_price: Any) -> list[dict[str, Any]]:
    points: list[dict[str, Any]] = []
    for row in price_rows:
        close = rounded(row.get("close"), 4)
        price_date = clean_text(row.get("price_date"))
        offset = clean_text(row.get("trading_day_offset"))
        if offset.upper() in {"T0", "0"}:
            continue
        if close is None and not price_date:
            continue
        points.append(
            {
                "trading_day_offset": offset,
                "price_date": price_date,
                "close": close,
                "return_pct": pct_change(close, base_price),
            }
        )
    return sorted(
        points,
        key=lambda point: (
            clean_text(point.get("price_date")),
            offset_sort_key(point.get("trading_day_offset")),
        ),
    )


def review_from_performance(
    row: dict[str, Any] | None,
    price_points: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    points = price_points or []
    if not row:
        return {
            "status": "missing_review",
            "result_label": "",
            "returns": {},
            "latest_price": None,
            "latest_price_date": "",
            "price_points": points,
            "failure_reason": "",
            "data_status": "",
        }
    returns = {field: rounded(row.get(field)) for field in RETURN_FIELDS}
    has_return = any(value is not None for value in returns.values())
    return {
        "status": "reviewed" if has_return else "insufficient_data",
        "result_label": clean_text(row.get("result_label")),
        "returns": returns,
        "latest_price": rounded(row.get("latest_price"), 4),
        "latest_price_date": clean_text(row.get("latest_price_date")),
        "price_points": points,
        "max_gain_3d_pct": rounded(row.get("max_gain_3d_pct")),
        "max_drawdown_3d_pct": rounded(row.get("max_drawdown_3d_pct")),
        "hit_stop_loss": as_bool(row.get("hit_stop_loss")),
        "hit_take_profit": as_bool(row.get("hit_take_profit")),
        "failure_reason": clean_text(row.get("failure_reason")),
        "data_status": clean_text(row.get("data_status")),
    }


def build_pick_from_score(row: dict[str, Any], rank: int, performance: dict[str, Any] | None = None) -> dict[str, Any]:
    symbol = clean_text(row.get("symbol") or row.get("stock_code"))
    stock_code = normalize_stock_code(symbol)
    return {
        "rank": as_int(row.get("rank"), rank),
        "symbol": symbol_from_code(symbol),
        "stock_code": stock_code,
        "name": first_present(row, "name", "stock_name"),
        "sector": first_present(row, "sector", "industry"),
        "total_score": compact_number(row.get("total_score")),
        "score_breakdown": {
            "trend": compact_number(row.get("trend_score")),
            "startup": compact_number(row.get("startup_score")),
            "sector": compact_number(row.get("sector_score")),
            "market": compact_number(row.get("market_score")),
        },
        "decision": first_present(row, "decision", "participation_level", "suggested_action"),
        "continuation": clean_text(row.get("continuation")),
        "buy_model": first_present(row, "buy_model", "strategy_label"),
        "notes": first_present(row, "notes", "selection_reason"),
        "risks": first_present(row, "risks", "risk_note", "hard_rejects"),
        "hard_rejects": clean_text(row.get("hard_rejects")),
        "plan": clean_text(row.get("plan")),
        "selection_price": rounded(row.get("selection_price"), 4),
        "stop_loss_price": rounded(row.get("stop_loss_price"), 4),
        "take_profit_price": rounded(row.get("take_profit_price"), 4),
        "review": review_from_performance(performance),
    }


def build_pick_from_selected(
    row: dict[str, Any],
    rank: int,
    performance: dict[str, Any] | None = None,
    price_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    stock_code = normalize_stock_code(row.get("stock_code"))
    base_price = (performance or {}).get("selection_price") or row.get("selection_price")
    return {
        "rank": as_int(row.get("rank_in_run"), rank),
        "symbol": symbol_from_code(stock_code),
        "stock_code": stock_code,
        "name": clean_text(row.get("stock_name")),
        "sector": first_present(row, "sector", "industry"),
        "total_score": compact_number(row.get("total_score")),
        "score_breakdown": {"trend": None, "startup": None, "sector": None, "market": None},
        "decision": first_present(row, "participation_level", "suggested_action"),
        "continuation": "",
        "buy_model": clean_text(row.get("strategy_label")),
        "notes": clean_text(row.get("selection_reason")),
        "risks": clean_text(row.get("risk_note")),
        "hard_rejects": "",
        "plan": clean_text(row.get("suggested_action")),
        "selection_price": rounded(row.get("selection_price"), 4),
        "stop_loss_price": rounded(row.get("stop_loss_price"), 4),
        "take_profit_price": rounded(row.get("take_profit_price"), 4),
        "review": review_from_performance(performance, build_price_points(price_rows or [], base_price)),
    }


def counter_options(values: Iterable[str]) -> list[dict[str, Any]]:
    counter = Counter(value for value in values if value)
    return [
        {"value": value, "count": count}
        for value, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]


def build_filters(picks: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    return {
        "decisions": counter_options(clean_text(pick.get("decision")) for pick in picks),
        "sectors": counter_options(clean_text(pick.get("sector")) for pick in picks),
        "buy_models": counter_options(clean_text(pick.get("buy_model")) for pick in picks),
        "review_statuses": counter_options(clean_text(pick.get("review", {}).get("status")) for pick in picks),
    }


def average_score(picks: list[dict[str, Any]]) -> float | None:
    values = [as_float(pick.get("total_score")) for pick in picks if as_float(pick.get("total_score")) is not None]
    return round(statistics.fmean(values), 2) if values else None


def top_score(picks: list[dict[str, Any]]) -> int | float | None:
    values = [as_float(pick.get("total_score")) for pick in picks if as_float(pick.get("total_score")) is not None]
    if not values:
        return None
    value = max(values)
    return int(value) if value.is_integer() else round(value, 2)


def score_buckets(picks: list[dict[str, Any]]) -> dict[str, int]:
    buckets = {"strong": 0, "trial": 0, "watch": 0, "avoid": 0, "unknown": 0}
    for pick in picks:
        score = as_float(pick.get("total_score"))
        if score is None:
            buckets["unknown"] += 1
        elif score >= 80:
            buckets["strong"] += 1
        elif score >= 65:
            buckets["trial"] += 1
        elif score >= 50:
            buckets["watch"] += 1
        else:
            buckets["avoid"] += 1
    return buckets


def build_review_summary(summary: dict[str, Any] | None, performance_rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not performance_rows:
        return {
            "status": "missing_review",
            "has_review": False,
            "empty_state": {
                "title": "暂无复盘数据",
                "message": "这次选股还没有写入后续价格或分析结果。",
            },
            "summary": {},
        }
    has_any_return = any(as_float(row.get(field)) is not None for row in performance_rows for field in RETURN_FIELDS)
    has_t3_or_later = any(
        as_float(row.get(field)) is not None
        for row in performance_rows
        for field in ("return_t3_close_pct", "return_t5_close_pct", "return_t10_close_pct")
    )
    if has_t3_or_later or as_int((summary or {}).get("valid_stock_count"), 0) > 0:
        status = "ready"
    elif has_any_return:
        status = "partial_review"
    else:
        status = "insufficient_data"
    empty_state = None
    if status == "insufficient_data":
        empty_state = {
            "title": "复盘数据不足",
            "message": "已有复盘记录，但后续价格不足，暂时不能判断胜率。",
        }
    summary = summary or {}
    return {
        "status": status,
        "has_review": True,
        "empty_state": empty_state,
        "summary": {
            "valid_stock_count": as_int(summary.get("valid_stock_count"), 0),
            "win_rate_t1": rounded(summary.get("win_rate_t1")),
            "win_rate_t2": rounded(summary.get("win_rate_t2")),
            "win_rate_t3": rounded(summary.get("win_rate_t3")),
            "win_rate_t5": rounded(summary.get("win_rate_t5")),
            "avg_return_latest_pct": rounded(summary.get("avg_return_latest_pct")),
            "avg_return_t1_pct": rounded(summary.get("avg_return_t1_pct")),
            "avg_return_t2_pct": rounded(summary.get("avg_return_t2_pct")),
            "avg_return_t3_pct": rounded(summary.get("avg_return_t3_pct")),
            "avg_return_t5_pct": rounded(summary.get("avg_return_t5_pct")),
            "avg_max_drawdown_3d_pct": rounded(summary.get("avg_max_drawdown_3d_pct")),
            "hit_stop_loss_count": as_int(summary.get("hit_stop_loss_count"), 0),
            "hit_take_profit_count": as_int(summary.get("hit_take_profit_count"), 0),
            "best_stock": clean_text(summary.get("best_stock")),
            "worst_stock": clean_text(summary.get("worst_stock")),
            "conclusion": clean_text(summary.get("conclusion")),
        },
    }


def build_metrics(picks: list[dict[str, Any]], review: dict[str, Any]) -> dict[str, Any]:
    decisions = Counter(clean_text(pick.get("decision")) or "未填写" for pick in picks)
    return {
        "selected_count": len(picks),
        "top_score": top_score(picks),
        "average_score": average_score(picks),
        "sector_count": len({clean_text(pick.get("sector")) for pick in picks if clean_text(pick.get("sector"))}),
        "decision_counts": dict(sorted(decisions.items())),
        "score_buckets": score_buckets(picks),
        "review_status": review["status"],
        "review_has_data": review["has_review"],
        "review": review["summary"],
    }


def weighted_average_from_summaries(
    summaries: list[dict[str, Any]],
    value_key: str,
    weight_key: str = "valid_stock_count",
) -> float | None:
    weighted_total = 0.0
    total_weight = 0
    for summary in summaries:
        value = as_float(summary.get(value_key))
        weight = as_int(summary.get(weight_key), 0)
        if value is None or weight <= 0:
            continue
        weighted_total += value * weight
        total_weight += weight
    return None if total_weight <= 0 else round(weighted_total / total_weight, 2)


def average_values(values: list[float]) -> float | None:
    return round(sum(values) / len(values), 2) if values else None


def win_rate(values: list[float]) -> float | None:
    return round(sum(1 for value in values if value > 0) / len(values) * 100, 2) if values else None


def build_strategy_effectiveness(run_payloads: list[dict[str, Any]]) -> dict[str, Any]:
    reviewed = [payload for payload in run_payloads if payload.get("review", {}).get("has_review")]
    summaries = [payload.get("metrics", {}).get("review", {}) for payload in reviewed]
    latest_values: list[float] = []
    t3_values: list[float] = []
    coalesced_values: list[float] = []
    for payload in reviewed:
        for pick in payload.get("picks", []):
            returns = pick.get("review", {}).get("returns", {})
            latest_value = as_float(returns.get("return_latest_pct"))
            t3_value = as_float(returns.get("return_t3_close_pct"))
            if latest_value is not None:
                latest_values.append(latest_value)
            if t3_value is not None:
                t3_values.append(t3_value)
            if t3_value is not None:
                coalesced_values.append(t3_value)
            elif latest_value is not None:
                coalesced_values.append(latest_value)
    valid_stock_count = len(coalesced_values) or sum(as_int(summary.get("valid_stock_count"), 0) for summary in summaries)
    t3_win_rate = win_rate(t3_values) or weighted_average_from_summaries(summaries, "win_rate_t3")
    latest_win_rate = win_rate(latest_values)
    avg_latest = average_values(latest_values) or weighted_average_from_summaries(summaries, "avg_return_latest_pct")
    avg_t3 = average_values(t3_values) or weighted_average_from_summaries(summaries, "avg_return_t3_pct")
    avg_drawdown = weighted_average_from_summaries(summaries, "avg_max_drawdown_3d_pct")
    stop_loss_count = sum(as_int(summary.get("hit_stop_loss_count"), 0) for summary in summaries)
    take_profit_count = sum(as_int(summary.get("hit_take_profit_count"), 0) for summary in summaries)
    if valid_stock_count <= 0:
        conclusion = "样本不足"
    elif (avg_t3 or 0) > 0 and (t3_win_rate or 0) >= 50:
        conclusion = "正向验证"
    elif (avg_t3 or 0) > 0:
        conclusion = "小样本正收益"
    elif not t3_values and (avg_latest or 0) > 0:
        conclusion = "最新价正向，等待 T3"
    elif not t3_values:
        conclusion = "最新价回撤，等待 T3"
    else:
        conclusion = "需要优化"
    return {
        "run_count": len(run_payloads),
        "reviewed_run_count": len(reviewed),
        "valid_stock_count": valid_stock_count,
        "win_rate_t3": t3_win_rate,
        "win_rate_latest": latest_win_rate,
        "avg_return_latest_pct": avg_latest,
        "avg_return_t3_pct": avg_t3,
        "avg_max_drawdown_3d_pct": avg_drawdown,
        "hit_stop_loss_count": stop_loss_count,
        "hit_take_profit_count": take_profit_count,
        "conclusion": conclusion,
    }


def build_run_payload(
    run: dict[str, Any],
    picks: list[dict[str, Any]],
    summary: dict[str, Any] | None,
    performance_rows: list[dict[str, Any]],
    source: dict[str, Any],
    generated_at: str,
) -> dict[str, Any]:
    run_id = clean_text(run.get("run_id"))
    date_key = date_key_from_text(run.get("selection_date") or run_id)
    review = build_review_summary(summary, performance_rows)
    metrics = build_metrics(picks, review)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "date": date_key,
        "selection_date": display_date(date_key),
        "run": {
            "run_id": run_id,
            "selection_date": display_date(date_key),
            "selection_time": clean_text(run.get("selection_time")),
            "strategy_version": clean_text(run.get("strategy_version")) or "v1_0",
            "market_env": clean_text(run.get("market_env")),
            "total_selected_count": as_int(run.get("total_selected_count"), len(picks)),
            "data_source": clean_text(run.get("data_source")),
            "operator": clean_text(run.get("operator")),
            "notes": clean_text(run.get("notes")),
            "source_type": clean_text(source.get("type")),
        },
        "metrics": metrics,
        "filters": build_filters(picks),
        "review": review,
        "picks": sorted(picks, key=lambda pick: as_int(pick.get("rank"), 999999)),
        "source": source,
    }


def build_runs_from_workbook(records: dict[str, list[dict[str, Any]]], generated_at: str) -> list[dict[str, Any]]:
    selected_by_run = group_by(records.get("selected_stocks", []), "run_id")
    performance_by_run = group_by(records.get("performance", []), "run_id")
    future_prices_by_run = group_by(records.get("future_prices", []), "run_id")
    summary_by_run = {clean_text(row.get("run_id")): row for row in records.get("summary_by_run", [])}
    result: list[dict[str, Any]] = []
    for run in records.get("selection_runs", []):
        run_id = clean_text(run.get("run_id"))
        performance_rows = performance_by_run.get(run_id, [])
        performance_by_code = {normalize_stock_code(row.get("stock_code")): row for row in performance_rows}
        price_rows_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for price_row in future_prices_by_run.get(run_id, []):
            price_rows_by_code[normalize_stock_code(price_row.get("stock_code"))].append(price_row)
        picks = [
            build_pick_from_selected(
                row,
                index,
                performance_by_code.get(normalize_stock_code(row.get("stock_code"))),
                price_rows_by_code.get(normalize_stock_code(row.get("stock_code")), []),
            )
            for index, row in enumerate(selected_by_run.get(run_id, []), start=1)
        ]
        if picks:
            result.append(
                build_run_payload(
                    run,
                    picks,
                    summary_by_run.get(run_id),
                    performance_rows,
                    {"type": "validation_workbook"},
                    generated_at,
                )
            )
    return result


def build_runs_from_scores(scores_dir: Path, referenced_paths: set[Path], generated_at: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for path in discover_score_files(scores_dir, referenced_paths):
        rows = read_csv_rows(path)
        if not rows:
            continue
        date_key = score_file_date(path)
        run = {
            "run_id": f"{date_key}_{path.stem}",
            "selection_date": display_date(date_key),
            "selection_time": "",
            "strategy_version": "v1_0",
            "market_env": "",
            "total_selected_count": len(rows),
            "data_source": str(path),
        }
        picks = [build_pick_from_score(row, index) for index, row in enumerate(rows, start=1)]
        result.append(build_run_payload(run, picks, None, [], {"type": "score_csv", "path": str(path)}, generated_at))
    return result


def active_run_sort_key(payload: dict[str, Any]) -> tuple[int, int, str, str]:
    review_rank = {"ready": 3, "partial_review": 2, "insufficient_data": 1, "missing_review": 0}
    run = payload["run"]
    return (
        as_int(run.get("total_selected_count"), len(payload.get("picks", []))),
        review_rank.get(clean_text(payload.get("review", {}).get("status")), 0),
        clean_text(run.get("selection_time")),
        clean_text(run.get("run_id")),
    )


def index_entry(date_key: str, active: dict[str, Any], run_count: int) -> dict[str, Any]:
    run = active["run"]
    metrics = active["metrics"]
    review = active["review"]
    return {
        "date": date_key,
        "selection_date": display_date(date_key),
        "run_id": run["run_id"],
        "label": f"{display_date(date_key)} {run.get('market_env') or run.get('strategy_version') or ''}".strip(),
        "market_env": run.get("market_env", ""),
        "total_selected_count": metrics["selected_count"],
        "top_score": metrics["top_score"],
        "average_score": metrics["average_score"],
        "score_buckets": metrics["score_buckets"],
        "has_review": review["has_review"],
        "review_status": review["status"],
        "run_count": run_count,
        "json_path": f"runs/{date_key}.json",
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_dashboard(
    scores_dir: Path = DEFAULT_SCORES_DIR,
    validation_workbook: Path = DEFAULT_VALIDATION_WORKBOOK,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    project_root: Path = PROJECT_ROOT,
) -> dict[str, Any]:
    generated_at = now_iso()
    workbook_records = read_workbook_records(validation_workbook)
    workbook_runs = build_runs_from_workbook(workbook_records, generated_at)
    referenced_paths = extract_referenced_score_paths(workbook_records.get("selection_runs", []), project_root)
    score_runs = build_runs_from_scores(scores_dir, referenced_paths, generated_at)
    runs = workbook_runs + score_runs
    strategy_effectiveness = build_strategy_effectiveness(runs)

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for payload in runs:
        date_key = clean_text(payload.get("date"))
        if date_key:
            grouped[date_key].append(payload)

    index_rows: list[dict[str, Any]] = []
    runs_dir = output_dir / "runs"
    for date_key, date_runs in sorted(grouped.items(), reverse=True):
        active = max(date_runs, key=active_run_sort_key)
        detail = {
            "schema_version": SCHEMA_VERSION,
            "generated_at": generated_at,
            "date": date_key,
            "selection_date": display_date(date_key),
            "active_run_id": active["run"]["run_id"],
            "run_choices": [
                {
                    "run_id": item["run"]["run_id"],
                    "selection_time": item["run"].get("selection_time", ""),
                    "market_env": item["run"].get("market_env", ""),
                    "total_selected_count": item["metrics"].get("selected_count", 0),
                    "review_status": item["review"].get("status", "missing_review"),
                }
                for item in sorted(date_runs, key=active_run_sort_key, reverse=True)
            ],
            "run": active["run"],
            "metrics": active["metrics"],
            "strategy_effectiveness": strategy_effectiveness,
            "filters": active["filters"],
            "review": active["review"],
            "picks": active["picks"],
            "source": active["source"],
        }
        write_json(runs_dir / f"{date_key}.json", detail)
        index_rows.append(index_entry(date_key, active, len(date_runs)))

    index_rows.sort(key=lambda row: row["date"], reverse=True)
    index_payload = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "latest_date": index_rows[0]["date"] if index_rows else "",
        "runs": index_rows,
        "strategy_effectiveness": strategy_effectiveness,
        "source": {
            "scores_dir": str(scores_dir),
            "validation_workbook": str(validation_workbook),
            "run_count": len(index_rows),
        },
    }
    write_json(output_dir / "runs_index.json", index_payload)
    return index_payload


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build local JSON data for the stock selection dashboard.")
    parser.add_argument("--scores-dir", type=Path, default=DEFAULT_SCORES_DIR)
    parser.add_argument("--validation-workbook", type=Path, default=DEFAULT_VALIDATION_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    index = build_dashboard(args.scores_dir, args.validation_workbook, args.output_dir)
    print(f"Dashboard runs: {len(index['runs'])}")
    print(f"Index JSON: {args.output_dir / 'runs_index.json'}")
    print(f"Run JSON dir: {args.output_dir / 'runs'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
