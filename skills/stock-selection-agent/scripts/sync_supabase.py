from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

DEFAULT_SCORES = PROJECT_ROOT / "outputs" / "selection_scores.csv"
DEFAULT_CANDIDATES = PROJECT_ROOT / "data" / "snapshots"
DEFAULT_WORKBOOK = PROJECT_ROOT / "data" / "validation" / "output" / "stock_selection_log.xlsx"
DEFAULT_REPORT = PROJECT_ROOT / "outputs" / "selection_report.md"

TABLE_ORDER = [
    "stock_selection_runs",
    "stock_selection_results",
    "stock_selection_prices",
    "stock_selection_performance",
]

UPSERT_CONFLICTS = {
    "stock_selection_runs": "run_id",
    "stock_selection_results": "run_id,stock_code",
    "stock_selection_prices": "run_id,stock_code,trading_day_offset",
    "stock_selection_performance": "run_id,stock_code",
}

TABLE_COLUMNS = {
    "stock_selection_runs": [
        "run_id",
        "selection_date",
        "selection_time",
        "strategy_version",
        "market_env",
        "total_selected_count",
        "data_source",
        "source_provider",
        "source_snapshot",
        "source_metadata",
        "dashboard",
        "report_markdown",
        "operator",
        "notes",
        "generated_at",
        "created_at",
        "updated_at",
    ],
    "stock_selection_results": [
        "run_id",
        "stock_code",
        "symbol",
        "stock_name",
        "sector",
        "rank_in_run",
        "total_score",
        "trend_score",
        "startup_score",
        "sector_score",
        "market_score",
        "decision",
        "continuation",
        "buy_model",
        "notes",
        "risks",
        "hard_rejects",
        "plan",
        "selection_price",
        "stop_loss_price",
        "take_profit_price",
        "candidate_payload",
        "score_payload",
        "created_at",
        "updated_at",
    ],
    "stock_selection_prices": [
        "run_id",
        "stock_code",
        "stock_name",
        "selection_date",
        "trading_day_offset",
        "price_date",
        "open",
        "high",
        "low",
        "close",
        "volume",
        "amount",
        "turnover_rate",
        "is_suspended",
        "data_source",
        "price_payload",
        "created_at",
        "updated_at",
    ],
    "stock_selection_performance": [
        "run_id",
        "stock_code",
        "stock_name",
        "selection_date",
        "sector",
        "industry",
        "concept",
        "strategy_label",
        "participation_level",
        "total_score",
        "rank_in_run",
        "selection_price",
        "latest_price",
        "latest_price_date",
        "return_latest_pct",
        "return_t1_close_pct",
        "return_t2_close_pct",
        "return_t3_close_pct",
        "return_t5_close_pct",
        "return_t10_close_pct",
        "max_high_3d",
        "max_gain_3d_pct",
        "min_low_3d",
        "max_drawdown_3d_pct",
        "max_high_5d",
        "max_gain_5d_pct",
        "min_low_5d",
        "max_drawdown_5d_pct",
        "hit_stop_loss",
        "hit_take_profit",
        "uncertain_intraday_order",
        "is_profitable_latest",
        "is_profitable_t1",
        "is_profitable_t2",
        "is_profitable_t3",
        "is_profitable_t5",
        "result_label",
        "failure_reason",
        "data_status",
        "performance_payload",
        "created_at",
        "updated_at",
    ],
}

COMPACT_SQL_EXCLUDED_COLUMNS = {
    "stock_selection_runs": {"dashboard", "report_markdown"},
    "stock_selection_results": {"candidate_payload", "score_payload", "risks", "hard_rejects"},
    "stock_selection_prices": {"price_payload"},
    "stock_selection_performance": {"performance_payload"},
}

REQUIRED_ENV = ("SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY")
DEFAULT_ENV_FILE = PROJECT_ROOT / "config" / "local.env"


def parse_env_file(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value
    return values


def load_sync_env(env_file: Path | None, base_env: dict[str, str] | None = None) -> dict[str, str]:
    merged = dict(base_env if base_env is not None else os.environ)
    if env_file is None:
        return merged
    for key, value in parse_env_file(env_file).items():
        merged.setdefault(key, value)
    return merged


@dataclass
class SyncPayload:
    runs: list[dict[str, Any]]
    results: list[dict[str, Any]]
    prices: list[dict[str, Any]]
    performance: list[dict[str, Any]]

    def table_payloads(self) -> dict[str, list[dict[str, Any]]]:
        return {
            "stock_selection_runs": self.runs,
            "stock_selection_results": self.results,
            "stock_selection_prices": self.prices,
            "stock_selection_performance": self.performance,
        }

    def counts(self) -> dict[str, int]:
        return {table: len(rows) for table, rows in self.table_payloads().items()}


def read_csv_records(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def read_json(path: Path | None) -> dict[str, Any]:
    if not path or not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def read_text(path: Path | None) -> str | None:
    if not path or not path.exists():
        return None
    return path.read_text(encoding="utf-8-sig")


def read_workbook_records(path: Path | None) -> dict[str, list[dict[str, Any]]]:
    if not path or not path.exists():
        return {}
    workbook = load_workbook(path, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.max_row < 2:
            result[sheet_name] = []
            continue
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        rows: list[dict[str, Any]] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            record = {
                headers[idx]: serialise_value(values[idx]) if idx < len(values) else None
                for idx in range(len(headers))
                if headers[idx]
            }
            if any(value not in {None, ""} for value in record.values()):
                rows.append(record)
        result[sheet_name] = rows
    return result


def serialise_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        lowered = text.lower()
        if lowered in {"true", "yes", "y"}:
            return True
        if lowered in {"false", "no", "n"}:
            return False
        return text
    return value


def compact_record(row: dict[str, Any]) -> dict[str, Any]:
    return {key: clean_value(value) for key, value in row.items() if clean_value(value) is not None}


def align_bulk_rows(table: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return rows
    columns = [column for column in TABLE_COLUMNS[table] if any(column in row for row in rows)]
    return [{column: row.get(column) for column in columns} for row in rows]


def typed_payload_record(row: dict[str, Any]) -> dict[str, Any]:
    payload = compact_record(row)
    for key, value in list(payload.items()):
        lowered = key.lower()
        if lowered in {"rank", "total_score", "trend_score", "startup_score", "sector_score", "market_score"}:
            payload[key] = as_int(value)
        elif (
            lowered in {"close", "open", "high", "low", "volume", "amount", "turnover_rate"}
            or lowered.startswith("return_")
            or lowered.endswith("_pct")
            or lowered.endswith("_price")
            or lowered.endswith("_count")
            or lowered.endswith("_billion")
            or lowered.startswith("ma")
        ):
            payload[key] = as_float(value)
    return payload


def as_int(value: Any, default: int | None = None) -> int | None:
    value = clean_value(value)
    if value is None:
        return default
    try:
        return int(float(str(value).replace(",", "")))
    except ValueError:
        return default


def as_float(value: Any, default: float | None = None) -> float | None:
    value = clean_value(value)
    if value is None:
        return default
    try:
        return float(str(value).replace(",", "").replace("%", ""))
    except ValueError:
        return default


def as_bool(value: Any, default: bool | None = None) -> bool | None:
    value = clean_value(value)
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "pass", "ok"}:
        return True
    if text in {"0", "false", "no", "n", "fail"}:
        return False
    return default


def parse_date_text(value: Any) -> str | None:
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    return None


def parse_time_text(value: Any) -> str | None:
    value = clean_value(value)
    if value is None:
        return None
    match = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", str(value))
    if not match:
        return None
    hour, minute, second = match.groups()
    return f"{int(hour):02d}:{int(minute):02d}:{int(second or 0):02d}"


def parse_timestamp_text(value: Any) -> str | None:
    value = clean_value(value)
    if value is None:
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d{14}", text):
        return (
            f"{text[:4]}-{text[4:6]}-{text[6:8]}T"
            f"{text[8:10]}:{text[10:12]}:{text[12:14]}"
        )
    return text


def normalize_stock_code(value: Any) -> str:
    value = clean_value(value)
    if value is None:
        return ""
    text = str(value).strip().upper()
    if re.fullmatch(r"\d{6}\.(SH|SZ|BJ)", text):
        return text
    digits = re.sub(r"\D", "", text)
    if len(digits) < 6:
        return text
    code = digits[-6:]
    if code.startswith("6") or code.startswith("9"):
        suffix = "SH"
    elif code.startswith(("0", "2", "3")):
        suffix = "SZ"
    elif code.startswith(("4", "8")):
        suffix = "BJ"
    else:
        suffix = "CN"
    return f"{code}.{suffix}"


def bare_symbol(stock_code: str) -> str:
    return stock_code.split(".", 1)[0] if stock_code else ""


def relative_path(path: Path | None) -> str | None:
    if not path:
        return None
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT.resolve()))
    except ValueError:
        return str(path)


def infer_latest_candidate_path(snapshot_dir: Path = DEFAULT_CANDIDATES) -> Path | None:
    if not snapshot_dir.exists():
        return None
    candidates = sorted(snapshot_dir.glob("*_candidates.csv"), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def infer_selection_date(args: argparse.Namespace, metadata: dict[str, Any], workbook_run: dict[str, Any] | None) -> str:
    explicit = parse_date_text(args.selection_date)
    if explicit:
        return explicit
    if workbook_run:
        workbook_date = parse_date_text(workbook_run.get("selection_date"))
        if workbook_date:
            return workbook_date
    meta_date = parse_date_text(metadata.get("end_date") or metadata.get("generated_at"))
    if meta_date:
        return meta_date
    candidates_path = resolve_optional_path(args.candidates)
    if candidates_path:
        path_date = parse_date_text(candidates_path.name)
        if path_date:
            return path_date
    return datetime.now().strftime("%Y-%m-%d")


def normalize_strategy_version(value: Any) -> str:
    text = str(clean_value(value) or "v1_0")
    return re.sub(r"[^0-9A-Za-z_]+", "_", text).strip("_") or "v1_0"


def derive_run_id(selection_date: str, strategy_version: str, metadata: dict[str, Any]) -> str:
    day = selection_date.replace("-", "")
    stamp = "000000"
    generated_at = parse_timestamp_text(metadata.get("generated_at"))
    if generated_at:
        match = re.search(r"T(\d{2}):?(\d{2}):?(\d{2})", generated_at)
        if match:
            stamp = "".join(match.groups())
    return f"{day}_{stamp}_{normalize_strategy_version(strategy_version)}"


def latest_workbook_run(workbook_rows: dict[str, list[dict[str, Any]]]) -> dict[str, Any] | None:
    runs = workbook_rows.get("selection_runs") or []
    if not runs:
        return None
    return sorted(runs, key=lambda row: str(row.get("created_at") or row.get("run_id") or ""))[-1]


def workbook_run_ids(workbook_rows: dict[str, list[dict[str, Any]]]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for row in workbook_rows.get("selection_runs") or []:
        run_id = str(clean_value(row.get("run_id")) or "")
        if run_id and run_id not in seen:
            result.append(run_id)
            seen.add(run_id)
    return result


def merge_payloads(payloads: list[SyncPayload]) -> SyncPayload:
    return SyncPayload(
        runs=[row for payload in payloads for row in payload.runs],
        results=[row for payload in payloads for row in payload.results],
        prices=[row for payload in payloads for row in payload.prices],
        performance=[row for payload in payloads for row in payload.performance],
    )


def resolve_optional_path(value: str | Path | None) -> Path | None:
    if value is None:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def rows_for_run(rows: list[dict[str, Any]], run_id: str) -> list[dict[str, Any]]:
    return [row for row in rows if str(row.get("run_id") or "") == run_id]


def candidate_lookup(candidates: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in candidates:
        code = normalize_stock_code(row.get("symbol") or row.get("stock_code") or row.get("code"))
        if code:
            result[code] = row
    return result


def dashboard_payload(
    workbook_rows: dict[str, list[dict[str, Any]]],
    dashboard_summary_path: Path | None,
    dashboard_detail_path: Path | None,
    run_id: str,
) -> dict[str, Any]:
    dashboard: dict[str, Any] = {}
    summary_rows = read_csv_records(dashboard_summary_path)
    detail_rows = read_csv_records(dashboard_detail_path)
    if summary_rows:
        dashboard["dashboard_summary"] = [compact_record(row) for row in summary_rows]
    if detail_rows:
        dashboard["dashboard_detail"] = [compact_record(row) for row in detail_rows]
    for sheet_name in [
        "summary_by_run",
        "summary_by_stock",
        "summary_by_sector",
        "summary_by_strategy_label",
        "summary_by_participation_level",
        "comparison",
    ]:
        rows = workbook_rows.get(sheet_name) or []
        if sheet_name == "summary_by_run":
            rows = rows_for_run(rows, run_id)
        if rows:
            dashboard[sheet_name] = [compact_record(row) for row in rows]
    return dashboard


def build_run_payload(
    args: argparse.Namespace,
    run_id: str,
    selection_date: str,
    scores: list[dict[str, Any]],
    metadata: dict[str, Any],
    workbook_run: dict[str, Any] | None,
    dashboard: dict[str, Any],
    report_markdown: str | None,
) -> dict[str, Any]:
    strategy_version = normalize_strategy_version(
        args.strategy_version or (workbook_run or {}).get("strategy_version") or "v1_0"
    )
    generated_at = parse_timestamp_text(metadata.get("generated_at") or (workbook_run or {}).get("created_at"))
    data_source = (workbook_run or {}).get("data_source")
    if not data_source:
        data_source = "; ".join(
            item
            for item in [
                f"scores={relative_path(resolve_optional_path(args.scores))}" if args.scores else None,
                f"candidates={relative_path(resolve_optional_path(args.candidates))}" if args.candidates else None,
                f"metadata={relative_path(resolve_optional_path(args.metadata))}" if args.metadata else None,
            ]
            if item
        )
    return compact_record(
        {
            "run_id": run_id,
            "selection_date": selection_date,
            "selection_time": parse_time_text((workbook_run or {}).get("selection_time")),
            "strategy_version": strategy_version,
            "market_env": args.market_env or (workbook_run or {}).get("market_env"),
            "total_selected_count": as_int((workbook_run or {}).get("total_selected_count"), len(scores)),
            "data_source": data_source,
            "source_provider": metadata.get("source") or (workbook_run or {}).get("source_type") or "local_artifact",
            "source_snapshot": relative_path(resolve_optional_path(args.candidates)),
            "source_metadata": metadata,
            "dashboard": dashboard,
            "report_markdown": report_markdown,
            "operator": args.operator or (workbook_run or {}).get("operator"),
            "notes": args.notes or (workbook_run or {}).get("notes"),
            "generated_at": generated_at,
            "created_at": (workbook_run or {}).get("created_at") or generated_at,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
    )


def build_result_payloads(
    run_id: str,
    scores: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    selected_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    synced_at = datetime.now().isoformat(timespec="seconds")
    lookup = candidate_lookup(candidates)
    selected_lookup = {normalize_stock_code(row.get("stock_code")): row for row in selected_rows}
    rows: list[dict[str, Any]] = []
    source_rows = scores or [
        {
            "rank": row.get("rank_in_run"),
            "symbol": row.get("stock_code"),
            "name": row.get("stock_name"),
            "sector": row.get("sector"),
            "total_score": row.get("total_score"),
            "decision": row.get("participation_level"),
            "buy_model": row.get("strategy_label"),
            "notes": row.get("selection_reason"),
            "risks": row.get("risk_note"),
            "plan": row.get("suggested_action"),
        }
        for row in selected_rows
    ]
    for fallback_rank, score in enumerate(source_rows, start=1):
        stock_code = normalize_stock_code(score.get("symbol") or score.get("stock_code"))
        if not stock_code:
            continue
        candidate = lookup.get(stock_code, {})
        selected = selected_lookup.get(stock_code, {})
        rows.append(
            compact_record(
                {
                    "run_id": run_id,
                    "stock_code": stock_code,
                    "symbol": bare_symbol(stock_code),
                    "stock_name": score.get("name") or selected.get("stock_name") or candidate.get("name") or stock_code,
                    "sector": score.get("sector") or selected.get("sector") or candidate.get("sector"),
                    "rank_in_run": as_int(score.get("rank"), fallback_rank),
                    "total_score": as_int(score.get("total_score") or selected.get("total_score")),
                    "trend_score": as_int(score.get("trend_score")),
                    "startup_score": as_int(score.get("startup_score")),
                    "sector_score": as_int(score.get("sector_score")),
                    "market_score": as_int(score.get("market_score")),
                    "decision": score.get("decision") or selected.get("participation_level"),
                    "continuation": score.get("continuation"),
                    "buy_model": score.get("buy_model") or selected.get("strategy_label"),
                    "notes": score.get("notes") or selected.get("selection_reason"),
                    "risks": score.get("risks") or selected.get("risk_note"),
                    "hard_rejects": score.get("hard_rejects"),
                    "plan": score.get("plan") or selected.get("suggested_action"),
                    "selection_price": as_float(
                        score.get("selection_price")
                        or selected.get("selection_price")
                        or selected.get("selection_close")
                        or candidate.get("close")
                    ),
                    "stop_loss_price": as_float(score.get("stop_loss_price") or selected.get("stop_loss_price")),
                    "take_profit_price": as_float(score.get("take_profit_price") or selected.get("take_profit_price")),
                    "candidate_payload": typed_payload_record(candidate),
                    "score_payload": typed_payload_record(score),
                    "created_at": selected.get("created_at") or selected.get("updated_at") or synced_at,
                    "updated_at": synced_at,
                }
            )
        )
    return rows


def build_price_payloads(run_id: str, price_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    synced_at = datetime.now().isoformat(timespec="seconds")
    for row in price_rows:
        stock_code = normalize_stock_code(row.get("stock_code"))
        offset = clean_value(row.get("trading_day_offset"))
        if not stock_code or offset is None:
            continue
        rows.append(
            compact_record(
                {
                    "run_id": run_id,
                    "stock_code": stock_code,
                    "stock_name": row.get("stock_name"),
                    "selection_date": parse_date_text(row.get("selection_date")),
                    "trading_day_offset": str(offset),
                    "price_date": parse_date_text(row.get("price_date")),
                    "open": as_float(row.get("open")),
                    "high": as_float(row.get("high")),
                    "low": as_float(row.get("low")),
                    "close": as_float(row.get("close")),
                    "volume": as_float(row.get("volume")),
                    "amount": as_float(row.get("amount")),
                    "turnover_rate": as_float(row.get("turnover_rate")),
                    "is_suspended": as_bool(row.get("is_suspended"), False),
                    "data_source": row.get("data_source"),
                    "price_payload": typed_payload_record(row),
                    "created_at": row.get("created_at") or row.get("updated_at") or synced_at,
                    "updated_at": row.get("updated_at") or synced_at,
                }
            )
        )
    return rows


def build_performance_payloads(run_id: str, performance_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    synced_at = datetime.now().isoformat(timespec="seconds")
    numeric_fields = {
        "total_score",
        "rank_in_run",
        "selection_price",
        "latest_price",
        "return_latest_pct",
        "return_t1_close_pct",
        "return_t2_close_pct",
        "return_t3_close_pct",
        "return_t5_close_pct",
        "return_t10_close_pct",
        "max_high_3d",
        "max_gain_3d_pct",
        "min_low_3d",
        "max_drawdown_3d_pct",
        "max_high_5d",
        "max_gain_5d_pct",
        "min_low_5d",
        "max_drawdown_5d_pct",
    }
    bool_fields = {
        "hit_stop_loss",
        "hit_take_profit",
        "uncertain_intraday_order",
        "is_profitable_latest",
        "is_profitable_t1",
        "is_profitable_t2",
        "is_profitable_t3",
        "is_profitable_t5",
    }
    direct_fields = {
        "stock_name",
        "sector",
        "industry",
        "concept",
        "strategy_label",
        "participation_level",
        "result_label",
        "failure_reason",
        "data_status",
    }
    for row in performance_rows:
        stock_code = normalize_stock_code(row.get("stock_code"))
        if not stock_code:
            continue
        payload: dict[str, Any] = {
            "run_id": run_id,
            "stock_code": stock_code,
            "selection_date": parse_date_text(row.get("selection_date")),
            "latest_price_date": parse_date_text(row.get("latest_price_date")),
            "performance_payload": typed_payload_record(row),
            "created_at": row.get("created_at") or row.get("updated_at") or synced_at,
            "updated_at": row.get("updated_at") or synced_at,
        }
        for field in direct_fields:
            payload[field] = row.get(field)
        for field in numeric_fields:
            converter = as_int if field in {"total_score", "rank_in_run"} else as_float
            payload[field] = converter(row.get(field))
        for field in bool_fields:
            payload[field] = as_bool(row.get(field))
        rows.append(compact_record(payload))
    return rows


def build_sync_payload(args: argparse.Namespace) -> SyncPayload:
    if args.candidates is None:
        latest_candidates = infer_latest_candidate_path()
        args.candidates = str(latest_candidates) if latest_candidates else None
    scores_path = resolve_optional_path(args.scores)
    candidates_path = resolve_optional_path(args.candidates)
    metadata_path = resolve_optional_path(args.metadata)
    workbook_path = resolve_optional_path(args.workbook)
    report_path = resolve_optional_path(args.report)
    dashboard_summary_path = resolve_optional_path(args.dashboard_summary)
    dashboard_detail_path = resolve_optional_path(args.dashboard_detail)

    scores = read_csv_records(scores_path)
    candidates = read_csv_records(candidates_path)
    metadata = read_json(metadata_path)
    workbook_rows = read_workbook_records(workbook_path)

    active_run_id = resolve_payload_run_id(args, workbook_rows, metadata)
    if getattr(args, "include_workbook_runs", False) and workbook_rows.get("selection_runs"):
        run_ids = workbook_run_ids(workbook_rows)
        if active_run_id and active_run_id not in run_ids:
            run_ids.append(active_run_id)
        use_artifacts_for_active_run = bool(args.run_id)
        return merge_payloads(
            [
                build_sync_payload_for_run(
                    args,
                    workbook_rows,
                    scores,
                    candidates,
                    metadata,
                    report_path,
                    dashboard_summary_path,
                    dashboard_detail_path,
                    run_id,
                    use_current_artifacts=(use_artifacts_for_active_run and run_id == active_run_id),
                )
                for run_id in run_ids
            ]
        )

    return build_sync_payload_for_run(
        args,
        workbook_rows,
        scores,
        candidates,
        metadata,
        report_path,
        dashboard_summary_path,
        dashboard_detail_path,
        active_run_id,
        use_current_artifacts=True,
    )


def resolve_payload_run_id(
    args: argparse.Namespace,
    workbook_rows: dict[str, list[dict[str, Any]]],
    metadata: dict[str, Any],
) -> str:
    workbook_run = None
    if args.run_id:
        workbook_run = next(
            (row for row in workbook_rows.get("selection_runs", []) if str(row.get("run_id")) == args.run_id),
            None,
        )
    else:
        workbook_run = latest_workbook_run(workbook_rows)
    selection_date = infer_selection_date(args, metadata, workbook_run)
    strategy_version = normalize_strategy_version(args.strategy_version or (workbook_run or {}).get("strategy_version"))
    run_id = args.run_id or (workbook_run or {}).get("run_id") or derive_run_id(selection_date, strategy_version, metadata)
    return str(run_id)


def build_sync_payload_for_run(
    args: argparse.Namespace,
    workbook_rows: dict[str, list[dict[str, Any]]],
    scores: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    metadata: dict[str, Any],
    report_path: Path | None,
    dashboard_summary_path: Path | None,
    dashboard_detail_path: Path | None,
    run_id: str,
    use_current_artifacts: bool,
) -> SyncPayload:
    workbook_run = next(
        (row for row in workbook_rows.get("selection_runs", []) if str(row.get("run_id")) == run_id),
        None,
    )
    run_args = argparse.Namespace(**vars(args))
    run_args.run_id = run_id
    if not use_current_artifacts:
        run_args.selection_date = None
        run_args.market_env = None
        run_args.operator = None
        run_args.notes = None
        run_args.strategy_version = (workbook_run or {}).get("strategy_version") or args.strategy_version
    run_metadata = metadata if use_current_artifacts else {}
    run_scores = scores if use_current_artifacts else []
    run_candidates = candidates if use_current_artifacts else []
    run_report = read_text(report_path) if use_current_artifacts else None
    run_dashboard_summary = dashboard_summary_path if use_current_artifacts else None
    run_dashboard_detail = dashboard_detail_path if use_current_artifacts else None

    selection_date = infer_selection_date(run_args, run_metadata, workbook_run)
    strategy_version = normalize_strategy_version(run_args.strategy_version or (workbook_run or {}).get("strategy_version"))
    selected_rows = rows_for_run(workbook_rows.get("selected_stocks", []), run_id)
    price_rows = rows_for_run(workbook_rows.get("future_prices", []), run_id)
    performance_rows = rows_for_run(workbook_rows.get("performance", []), run_id)
    dashboard = dashboard_payload(workbook_rows, run_dashboard_summary, run_dashboard_detail, run_id)
    runs = [
        build_run_payload(
            run_args,
            run_id,
            selection_date,
            run_scores,
            run_metadata,
            workbook_run,
            dashboard,
            run_report,
        )
    ]
    return SyncPayload(
        runs=runs,
        results=build_result_payloads(run_id, run_scores, run_candidates, selected_rows),
        prices=build_price_payloads(run_id, price_rows),
        performance=build_performance_payloads(run_id, performance_rows),
    )


def execute_sync(payload: SyncPayload, dry_run: bool, env: dict[str, str] | None = None) -> dict[str, Any]:
    env = env if env is not None else os.environ
    summary = {
        "status": "dry_run" if dry_run else "pending",
        "counts": payload.counts(),
        "conflicts": UPSERT_CONFLICTS,
    }
    if dry_run:
        return summary
    missing = [name for name in REQUIRED_ENV if not env.get(name)]
    if missing:
        summary.update({"status": "skipped", "reason": f"missing env: {', '.join(missing)}"})
        return summary
    try:
        from supabase import create_client
        client = create_client(env["SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"])
    except (ImportError, ModuleNotFoundError, AttributeError):
        try:
            from backend.supabase_jobs import RestSupabaseClient
        except ModuleNotFoundError:
            summary.update({"status": "skipped", "reason": "Supabase client fallback is unavailable"})
            return summary
        client = RestSupabaseClient(env["SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"])
    written: dict[str, int] = {}
    for table in TABLE_ORDER:
        rows = align_bulk_rows(table, payload.table_payloads()[table])
        if not rows:
            written[table] = 0
            continue
        client.table(table).upsert(rows, on_conflict=UPSERT_CONFLICTS[table]).execute()
        written[table] = len(rows)
    summary.update({"status": "synced", "written": written})
    return summary


def sql_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def selected_columns(table: str, rows: list[dict[str, Any]], compact: bool = False) -> list[str]:
    available = {key for row in rows for key in row}
    if compact:
        available -= COMPACT_SQL_EXCLUDED_COLUMNS.get(table, set())
    return [column for column in TABLE_COLUMNS[table] if column in available]


def build_upsert_sql(table: str, rows: list[dict[str, Any]], compact: bool = False) -> str:
    if not rows:
        return f"-- no rows for public.{table}"
    columns = selected_columns(table, rows, compact=compact)
    conflicts = [item.strip() for item in UPSERT_CONFLICTS[table].split(",")]
    update_columns = [column for column in columns if column not in conflicts]
    column_sql = ", ".join(sql_identifier(column) for column in columns)
    conflict_sql = ", ".join(sql_identifier(column) for column in conflicts)
    if update_columns:
        update_sql = ", ".join(
            f"{sql_identifier(column)} = excluded.{sql_identifier(column)}"
            for column in update_columns
        )
    else:
        update_sql = f"{sql_identifier(conflicts[0])} = excluded.{sql_identifier(conflicts[0])}"
    payload_json = json.dumps([{column: row.get(column) for column in columns} for row in rows], ensure_ascii=True, default=str)
    tag = f"payload_{hashlib.sha1(payload_json.encode('utf-8')).hexdigest()[:12]}"
    return (
        f"with payload as (\n"
        f"    select * from jsonb_populate_recordset(null::public.{sql_identifier(table)}, ${tag}${payload_json}${tag}$::jsonb)\n"
        f")\n"
        f"insert into public.{sql_identifier(table)} ({column_sql})\n"
        f"select {column_sql} from payload\n"
        f"on conflict ({conflict_sql}) do update set\n"
        f"    {update_sql};"
    )


def chunk_rows(rows: list[dict[str, Any]], chunk_size: int) -> list[list[dict[str, Any]]]:
    if chunk_size <= 0:
        return [rows]
    return [rows[idx : idx + chunk_size] for idx in range(0, len(rows), chunk_size)]


def write_sql_bundle(
    payload: SyncPayload,
    output_dir: Path,
    chunk_size: int,
    compact: bool,
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    for stale in list(output_dir.glob("*.sql")) + [output_dir / "manifest.json"]:
        if stale.exists():
            stale.unlink()
    files: list[dict[str, Any]] = []
    sequence = 1
    for table in TABLE_ORDER:
        rows = payload.table_payloads()[table]
        chunks = chunk_rows(rows, chunk_size)
        if not chunks:
            continue
        for chunk_index, chunk in enumerate(chunks):
            filename = f"{sequence:03d}_{table}_{chunk_index:03d}.sql"
            path = output_dir / filename
            path.write_text(build_upsert_sql(table, chunk, compact=compact) + "\n", encoding="utf-8")
            files.append(
                {
                    "table": table,
                    "chunk_index": chunk_index,
                    "row_count": len(chunk),
                    "path": str(path),
                }
            )
            sequence += 1
    manifest = {
        "schema_version": 1,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "counts": payload.counts(),
        "conflicts": UPSERT_CONFLICTS,
        "chunk_size": chunk_size,
        "compact": compact,
        "files": files,
    }
    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    manifest["manifest_path"] = str(manifest_path)
    return manifest


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build and optionally upsert stock-selection payloads into Supabase.")
    parser.add_argument("--scores", default=str(DEFAULT_SCORES), help="Scored selection CSV.")
    parser.add_argument("--candidates", default=None, help="Candidate snapshot CSV. Defaults to latest data/snapshots/*_candidates.csv.")
    parser.add_argument("--metadata", default=None, help="Snapshot metadata JSON.")
    parser.add_argument("--report", default=str(DEFAULT_REPORT), help="Selection report Markdown.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Validation/performance workbook.")
    parser.add_argument("--dashboard-summary", default=None, help="Optional dashboard summary CSV.")
    parser.add_argument("--dashboard-detail", default=None, help="Optional dashboard detail CSV.")
    parser.add_argument("--run-id", default=None, help="Run id to sync. Defaults to latest workbook run or derived metadata run.")
    parser.add_argument(
        "--include-workbook-runs",
        action="store_true",
        help="Also sync every run in the validation workbook so historical price updates reach Supabase.",
    )
    parser.add_argument("--selection-date", default=None, help="Override selection date.")
    parser.add_argument("--strategy-version", default="v1_0", help="Strategy version label.")
    parser.add_argument("--market-env", default=None, help="Optional market environment label.")
    parser.add_argument("--operator", default=None, help="Optional operator label.")
    parser.add_argument("--notes", default=None, help="Optional run notes.")
    parser.add_argument("--dry-run", action="store_true", help="Build payloads but skip Supabase writes.")
    parser.add_argument("--print-payload", action="store_true", help="Print full payload JSON before the sync summary.")
    parser.add_argument("--emit-sql", action="store_true", help="Print one SQL upsert chunk for Supabase execute_sql.")
    parser.add_argument("--sql-table", choices=TABLE_ORDER, default="stock_selection_runs", help="Table to emit when --emit-sql is set.")
    parser.add_argument("--sql-chunk-size", type=int, default=50, help="Rows per emitted SQL chunk.")
    parser.add_argument("--sql-chunk-index", type=int, default=0, help="Zero-based SQL chunk index to print.")
    parser.add_argument("--sql-compact", action="store_true", help="Omit large redundant payload columns from emitted SQL.")
    parser.add_argument("--write-sql-dir", type=Path, help="Write all table upsert SQL chunks for Supabase MCP execution.")
    parser.add_argument(
        "--env-file",
        type=Path,
        default=DEFAULT_ENV_FILE,
        help="Local env file for Supabase server-side credentials. Defaults to config/local.env.",
    )
    parser.add_argument("--fail-on-skip", action="store_true", help="Exit non-zero when a non-dry-run sync is skipped.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    payload = build_sync_payload(args)
    if args.emit_sql:
        rows = payload.table_payloads()[args.sql_table]
        chunks = chunk_rows(rows, args.sql_chunk_size)
        if not chunks:
            print(f"-- no rows for {args.sql_table}")
            return 0
        if args.sql_chunk_index >= len(chunks):
            raise SystemExit(f"chunk index {args.sql_chunk_index} out of range; chunk_count={len(chunks)}")
        print(build_upsert_sql(args.sql_table, chunks[args.sql_chunk_index], compact=args.sql_compact))
        return 0
    if args.print_payload:
        print(json.dumps(payload.table_payloads(), ensure_ascii=False, indent=2, default=str))
    bundle = None
    if args.write_sql_dir:
        bundle = write_sql_bundle(payload, args.write_sql_dir, args.sql_chunk_size, args.sql_compact)
    result = execute_sync(payload, dry_run=args.dry_run, env=load_sync_env(args.env_file))
    if bundle:
        result["sql_bundle"] = bundle
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    if args.fail_on_skip and result.get("status") == "skipped":
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
